from aiogram import Router, F
from aiogram.types import BufferedInputFile, Message
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

router = Router()

ADMIN_ID = 388622523


class BroadcastForwardState(StatesGroup):
    waiting_post = State()


class WelcomeVoiceState(StatesGroup):
    waiting_voice = State()


class OnboardingVideoState(StatesGroup):
    waiting_video = State()


def is_admin(user_id: int) -> bool:
    return user_id == ADMIN_ID


PROMO_TIER_NAMES = {
    "scan_text": "Скан и текст",
    "base": "База",
    "premium": "Премиум",
    "business": "Бизнес",
}


def _command_args(message: Message, command: str) -> str:
    text = (message.text or "").strip()
    return text[len(command):].strip()


def _format_dt(value) -> str:
    if not value:
        return "нет"
    try:
        return value.strftime("%d.%m.%Y %H:%M")
    except AttributeError:
        return str(value)


def _format_date(value) -> str:
    if not value:
        return "нет"
    try:
        return value.strftime("%d.%m.%Y")
    except AttributeError:
        return str(value)


def _format_promo_limit(used_count: int, max_uses) -> str:
    if max_uses is None:
        return str(used_count) + "/без лимита"
    return str(used_count) + "/" + str(max_uses)


def _clean_promo_code(raw_code: str) -> str:
    import re
    code = (raw_code or "").strip().upper()
    code = re.sub(r"[^A-Z0-9_-]", "", code)
    return code[:32]


def _promo_user_label(username, full_name, user_id) -> str:
    name = full_name or "без имени"
    if username:
        return name + " (@" + username + ")"
    return name + " (" + str(user_id) + ")"


@router.message(Command("admin_stats"))
async def cmd_admin_stats(message: Message):
    if not is_admin(message.from_user.id):
        return

    from app.database import fetchall, fetchone

    total_row = await fetchone("SELECT COUNT(*) FROM users")
    total = total_row[0] if total_row else 0

    tier_rows = await fetchall(
        "SELECT subscription_tier, COUNT(*) FROM users GROUP BY subscription_tier ORDER BY COUNT(*) DESC"
    )

    text = "📊 Статистика пользователей\n\n"
    text += "Всего: " + str(total) + "\n\n"
    text += "По тарифам:\n"
    for tier, count in tier_rows:
        text += "  " + str(tier) + ": " + str(count) + "\n"

    await message.answer(text, parse_mode=None)


@router.message(Command("setwelcomevoice"))
async def cmd_set_welcome_voice(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(WelcomeVoiceState.waiting_voice)
    await message.answer(
        "Кидай сюда приветственное голосовое. Можно просто переслать его из твоей группы.\n\n"
        "Я сохраню его для /start. Без лишней возни, как мы любим."
    )


@router.message(WelcomeVoiceState.waiting_voice, F.voice)
async def msg_set_welcome_voice(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()

    from app.database import set_bot_setting

    await set_bot_setting("welcome_voice_file_id", message.voice.file_id)
    await message.answer(
        "Готово. Теперь /start начинается с твоего голосового.\n\n"
        "Пользователь услышит тебя первым, а потом уже я начну финансово приставать."
    )


@router.message(WelcomeVoiceState.waiting_voice)
async def msg_set_welcome_voice_wrong(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await message.answer("Мне нужно именно голосовое. Текстом харизму не прикрутим.")


@router.message(Command("setonboardingvideo"))
async def cmd_set_onboarding_video(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(OnboardingVideoState.waiting_video)
    await message.answer(
        "Кидай сюда видео для автоонбординга. Подпись к видео сохраню как текст под роликом.\n\n"
        "Пользователю оно отправится автоматически после первой операции или первой финцели."
    )


@router.message(OnboardingVideoState.waiting_video, F.video)
async def msg_set_onboarding_video(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()

    from app.database import set_bot_setting

    await set_bot_setting("onboarding_video_file_id", message.video.file_id)
    await set_bot_setting("onboarding_video_caption", message.caption or "")
    await message.answer(
        "Готово. Видео сохранено и будет отправляться автоматически после первого действия пользователя."
    )


@router.message(OnboardingVideoState.waiting_video)
async def msg_set_onboarding_video_wrong(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await message.answer("Нужно именно видео. Пришли ролик, а подпись добавь прямо к нему.")


@router.message(Command("setonboardingcaption"))
async def cmd_set_onboarding_caption(message: Message):
    if not is_admin(message.from_user.id):
        return

    caption = (message.text or "").strip()[len("/setonboardingcaption"):].strip()
    if not caption:
        await message.answer(
            "Напиши текст сразу после команды:\n\n"
            "/setonboardingcaption Твой текст под видео"
        )
        return
    if len(caption) > 1024:
        await message.answer("Подпись к видео в Telegram максимум 1024 символа. Сократи чуть-чуть.")
        return

    from app.database import set_bot_setting

    await set_bot_setting("onboarding_video_caption", caption)
    await message.answer("Готово. Подпись к onboarding-видео обновлена.")


@router.message(Command("promo_create"))
async def cmd_promo_create(message: Message):
    if not is_admin(message.from_user.id):
        return

    import datetime as dt

    args = _command_args(message, "/promo_create").split()
    if len(args) < 4:
        await message.answer(
            "Формат:\n"
            "/promo_create CODE premium 90 100\n\n"
            "Можно с датой окончания:\n"
            "/promo_create CODE premium 90 100 2026-12-31\n\n"
            "Для безлимита вместо 100 напиши unlimited."
        )
        return

    code = _clean_promo_code(args[0])
    tier = args[1].strip().lower()
    if tier not in PROMO_TIER_NAMES:
        await message.answer("Тариф должен быть: scan_text, base, premium или business.")
        return

    try:
        days = int(args[2])
    except ValueError:
        await message.answer("Количество дней должно быть числом. Для 3 месяцев ставим 90.")
        return
    if days <= 0:
        await message.answer("Количество дней должно быть больше нуля.")
        return

    raw_limit = args[3].strip().lower()
    if raw_limit in {"unlimited", "infinite", "inf", "безлимит", "безлим"}:
        max_uses = None
    else:
        try:
            max_uses = int(raw_limit)
        except ValueError:
            await message.answer("Лимит использований должен быть числом или словом unlimited.")
            return
        if max_uses <= 0:
            await message.answer("Лимит использований должен быть больше нуля.")
            return

    expires_at = None
    if len(args) >= 5:
        try:
            expires_at = dt.datetime.strptime(args[4], "%Y-%m-%d")
        except ValueError:
            await message.answer("Дата окончания должна быть в формате YYYY-MM-DD, например 2026-12-31.")
            return

    if not code:
        await message.answer("Промокод пустой. Используй латиницу/цифры, например REFIVAN90.")
        return

    from app.database import fetchone

    row = await fetchone(
        """INSERT INTO promo_codes (code, tier, days, max_uses, expires_at)
           VALUES (%s, %s, %s, %s, %s)
           ON CONFLICT (code) DO UPDATE
           SET tier = EXCLUDED.tier,
               days = EXCLUDED.days,
               max_uses = EXCLUDED.max_uses,
               expires_at = EXCLUDED.expires_at
           RETURNING code, tier, days, max_uses, used_count, expires_at""",
        (code, tier, days, max_uses, expires_at),
    )

    await message.answer(
        "Готово, промокод настроен.\n\n"
        "Код: " + row[0] + "\n"
        "Тариф: " + PROMO_TIER_NAMES.get(row[1], row[1]) + "\n"
        "Срок: " + str(row[2]) + " дней\n"
        "Использования: " + _format_promo_limit(row[4], row[3]) + "\n"
        "Действует до: " + _format_date(row[5]) + "\n\n"
        "Проверить: /promo_check " + row[0] + "\n"
        "Excel: /promo_export " + row[0],
        parse_mode=None,
    )


@router.message(Command("promo_list"))
async def cmd_promo_list(message: Message):
    if not is_admin(message.from_user.id):
        return

    from app.database import fetchall

    rows = await fetchall(
        """SELECT code, tier, days, max_uses, used_count, expires_at, created_at
           FROM promo_codes
           ORDER BY created_at DESC, id DESC
           LIMIT 30"""
    )

    if not rows:
        await message.answer(
            "Промокодов пока нет.\n\n"
            "Создать 3 месяца премиума:\n"
            "/promo_create PREMIUM90 premium 90 100"
        )
        return

    text = "Промокоды: последние 30\n\n"
    for code, tier, days, max_uses, used_count, expires_at, created_at in rows:
        text += (
            code + " — " + PROMO_TIER_NAMES.get(tier, tier) + ", "
            + str(days) + " дн., "
            + _format_promo_limit(used_count, max_uses)
        )
        if expires_at:
            text += ", до " + _format_date(expires_at)
        text += "\n"

    text += "\nПроверить: /promo_check CODE\nExcel: /promo_export CODE"
    await message.answer(text, parse_mode=None)


@router.message(Command("promo_check"))
async def cmd_promo_check(message: Message):
    if not is_admin(message.from_user.id):
        return

    code = _clean_promo_code(_command_args(message, "/promo_check").split()[0] if _command_args(message, "/promo_check") else "")
    if not code:
        await message.answer("Формат: /promo_check CODE")
        return

    from app.database import fetchall, fetchone

    promo = await fetchone(
        """SELECT id, code, tier, days, max_uses, used_count, expires_at, created_at
           FROM promo_codes
           WHERE code = %s""",
        (code,),
    )
    if not promo:
        await message.answer("Промокод " + code + " не найден.")
        return

    promo_id, code, tier, days, max_uses, used_count, expires_at, created_at = promo
    uses = await fetchall(
        """SELECT u.id, u.username, u.full_name, pu.used_at
           FROM promo_uses pu
           JOIN users u ON u.id = pu.user_id
           WHERE pu.promo_id = %s
           ORDER BY pu.used_at DESC
           LIMIT 10""",
        (promo_id,),
    )

    text = (
        "Промокод " + code + "\n\n"
        "Тариф: " + PROMO_TIER_NAMES.get(tier, tier) + "\n"
        "Срок: " + str(days) + " дней\n"
        "Использования: " + _format_promo_limit(used_count, max_uses) + "\n"
        "Создан: " + _format_dt(created_at) + "\n"
        "Действует до: " + _format_date(expires_at) + "\n\n"
    )

    if uses:
        text += "Последние активации:\n"
        for index, (user_id, username, full_name, used_at) in enumerate(uses, start=1):
            text += (
                str(index) + ". "
                + _promo_user_label(username, full_name, user_id)
                + " — " + _format_dt(used_at) + "\n"
            )
        text += "\nExcel: /promo_export " + code
    else:
        text += "Активаций пока нет."

    await message.answer(text, parse_mode=None)


@router.message(Command("promo_export"))
async def cmd_promo_export(message: Message):
    if not is_admin(message.from_user.id):
        return

    code = _clean_promo_code(_command_args(message, "/promo_export").split()[0] if _command_args(message, "/promo_export") else "")
    if not code:
        await message.answer("Формат: /promo_export CODE")
        return

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from app.database import fetchall, fetchone

    promo = await fetchone(
        """SELECT id, code, tier, days, max_uses, used_count, expires_at, created_at
           FROM promo_codes
           WHERE code = %s""",
        (code,),
    )
    if not promo:
        await message.answer("Промокод " + code + " не найден.")
        return

    promo_id, code, tier, days, max_uses, used_count, expires_at, created_at = promo
    rows = await fetchall(
        """SELECT pu.used_at,
                  u.id,
                  u.username,
                  u.full_name,
                  u.subscription_tier,
                  u.premium_until,
                  u.created_at
           FROM promo_uses pu
           JOIN users u ON u.id = pu.user_id
           WHERE pu.promo_id = %s
           ORDER BY pu.used_at DESC""",
        (promo_id,),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Активации"

    ws.append(["Промокод", code])
    ws.append(["Тариф", PROMO_TIER_NAMES.get(tier, tier)])
    ws.append(["Срок, дней", days])
    ws.append(["Использования", _format_promo_limit(used_count, max_uses)])
    ws.append(["Действует до", _format_date(expires_at)])
    ws.append([])

    headers = [
        "Дата применения",
        "Telegram ID",
        "Username",
        "Имя аккаунта",
        "Текущий тариф",
        "Подписка до",
        "Дата регистрации",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5597")

    for used_at, user_id, username, full_name, current_tier, premium_until, user_created_at in rows:
        ws.append([
            _format_dt(used_at),
            user_id,
            username or "",
            full_name or "",
            PROMO_TIER_NAMES.get(current_tier, current_tier or ""),
            _format_dt(premium_until),
            _format_dt(user_created_at),
        ])

    for col_index, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 2, 12), 42)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await message.answer_document(
        BufferedInputFile(buf.read(), filename="promo_" + code + ".xlsx"),
        caption=(
            "Выгрузка по промокоду " + code + "\n"
            "Активаций: " + str(len(rows))
        ),
    )


@router.message(Command("broadcast"))
async def cmd_broadcast(message: Message):
    if not is_admin(message.from_user.id):
        return

    text = message.text.strip()[len("/broadcast"):].strip()
    if not text:
        await message.answer(
            "Формат: /broadcast Текст сообщения\n\n"
            "Сообщение будет отправлено всем пользователям бота."
        )
        return

    from app.database import fetchall
    users = await fetchall("SELECT id FROM users")

    sent = 0
    failed = 0
    status_msg = await message.answer("Рассылка началась... 0/" + str(len(users)))

    for i, (user_id,) in enumerate(users):
        try:
            await message.bot.send_message(user_id, text, parse_mode=None)
            sent += 1
        except Exception:
            failed += 1

        if (i + 1) % 50 == 0:
            try:
                await status_msg.edit_text(
                    "Рассылка идёт... " + str(i + 1) + "/" + str(len(users))
                )
            except Exception:
                pass

    await status_msg.edit_text(
        "Рассылка завершена!\nОтправлено: " + str(sent) + "\nНе удалось: " + str(failed)
    )



@router.message(Command("broadcast_forward"))
async def cmd_broadcast_forward(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(BroadcastForwardState.waiting_post)
    await message.answer(
        "Пришли пост (текст, фото, видео — что угодно), который нужно разослать всем пользователям.\n\n"
        "Можно переслать сообщение из любого канала."
    )


@router.message(BroadcastForwardState.waiting_post)
async def msg_broadcast_forward_post(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.clear()

    from app.database import fetchall
    users = await fetchall("SELECT id FROM users")

    sent = 0
    failed = 0
    status_msg = await message.answer("Рассылка началась... 0/" + str(len(users)))

    for i, (user_id,) in enumerate(users):
        try:
            await message.bot.copy_message(
                chat_id=user_id,
                from_chat_id=message.chat.id,
                message_id=message.message_id,
            )
            sent += 1
        except Exception:
            failed += 1

        if (i + 1) % 50 == 0:
            try:
                await status_msg.edit_text(
                    "Рассылка идёт... " + str(i + 1) + "/" + str(len(users))
                )
            except Exception:
                pass

    await status_msg.edit_text(
        "Рассылка завершена!\nОтправлено: " + str(sent) + "\nНе удалось: " + str(failed)
    )
