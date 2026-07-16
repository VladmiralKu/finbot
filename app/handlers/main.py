from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import default_state
from aiogram.filters import StateFilter
from aiogram.fsm.state import State, StatesGroup
from datetime import date, datetime, timedelta
import asyncio
import re
from html import escape

from app.database import (
    execute, fetchone, fetchall,
    get_or_create_user, get_categories,
    get_monthly_summary, get_recent_transactions,
    get_category_breakdown, is_premium,
)
from app.keyboards import (
    MAIN_REPLY_KB,
    main_menu,
    manual_input_keyboard,
    onboarding_finish_keyboard,
    categories_keyboard,
    confirm_keyboard,
    premium_keyboard,
)
from app.services.transaction_service import create_transaction
from app.services.insights import build_first_transaction_insight, build_transaction_insight
from app.services.transaction_public_ids import (
    public_number_for_transaction,
    public_numbers_for_ids,
    resolve_transaction_reference,
)

router = Router()


class AddTransaction(StatesGroup):
    choosing_category = State()
    entering_amount   = State()
    entering_comment  = State()


class QuickCategoryRenameState(StatesGroup):
    waiting_new_name = State()


class BeautifulReportState(StatesGroup):
    waiting_confirm = State()


NO_SPEND_REPLIES = {
    "ничего",
    "ничего не тратил",
    "ничего не тратила",
    "пока ничего",
    "сегодня ничего",
    "нет",
}


def _is_no_spend_reply(text: str) -> bool:
    normalized = " ".join((text or "").lower().strip().split())
    return normalized in NO_SPEND_REPLIES


def _format_comment_inline(comment: str | None) -> str:
    value = (comment or "").strip()
    return f" | 💬 «{value}»" if value else ""


def _format_comment_block(comment: str | None) -> str:
    value = (comment or "").strip()
    return f"💬 Комментарий: «{value}»" if value else "💬 Комментарий: —"


def _format_comment_block_html(comment: str | None) -> str:
    value = (comment or "").strip()
    if not value:
        return "💬 <b>Комментарий:</b> —"
    return f"💬 <b>Комментарий:</b> <i>«{escape(value)}»</i>"


def _looks_like_edit_confusion(text: str) -> bool:
    lower = (text or "").lower()
    action_like = (
        "замени" in lower
        or "заменить" in lower
        or "запени" in lower
        or "помен" in lower
        or "смени" in lower
        or "сменить" in lower
        or "переимен" in lower
    )
    object_like = "категор" in lower or "стать" in lower or "раздел" in lower
    return action_like and object_like


MANUAL_AMOUNT_RE = re.compile(r"(?<!\d)([+-]?\d[\d\s]*(?:[,.]\d{1,2})?)(?!\d)")
MANUAL_AMOUNT_WITH_CURRENCY_RE = re.compile(
    r"(?<!\d)([+-]?\d[\d\s]*(?:[,.]\d{1,2})?)\s*"
    r"(?:₽|р\.?|руб\.?|рублей|рубля|рубль)(?![а-яёa-z0-9])",
    re.IGNORECASE,
)
MANUAL_MONTH_MARKERS = (
    "январ", "феврал", "март", "апрел", "мая", "май", "июн",
    "июл", "август", "сентябр", "октябр", "ноябр", "декабр",
)


def _manual_amount_value(raw: str) -> float | None:
    try:
        return abs(float(raw.replace(" ", "").replace(",", ".").lstrip("+-")))
    except (TypeError, ValueError):
        return None


def _manual_transaction_date(text: str) -> date:
    lower = (text or "").lower()
    if "позавчера" in lower:
        return date.today() - timedelta(days=2)
    if "вчера" in lower:
        return date.today() - timedelta(days=1)
    return date.today()


def _parse_safe_manual_input(text: str) -> dict | None:
    from app.services.transaction_ai import _normalize_amount_phrases

    value = " ".join(_normalize_amount_phrases(text or "").strip().split())
    if not value:
        return None

    matches = list(MANUAL_AMOUNT_WITH_CURRENCY_RE.finditer(value))
    if not matches:
        matches = list(MANUAL_AMOUNT_RE.finditer(value))

    candidates = []
    lower = value.lower()
    for match in matches:
        amount = _manual_amount_value(match.group(1))
        if not amount or amount <= 0:
            continue
        if 1900 <= amount <= 2100 and any(marker in lower for marker in MANUAL_MONTH_MARKERS):
            continue
        candidates.append((amount, match))

    if not candidates:
        return None

    amount, match = max(candidates, key=lambda item: item[0])
    comment = (value[:match.start()] + " " + value[match.end():]).strip()
    comment = re.sub(r"\b(?:сегодня|вчера|позавчера)\b", " ", comment, flags=re.IGNORECASE)
    comment = re.sub(r"^\s*(?:за|на|в|по)\s+", "", comment, flags=re.IGNORECASE)
    comment = re.sub(r"\s+", " ", comment).strip(" .,!?:;-")
    return {
        "amount": amount,
        "comment": comment,
        "transaction_date": _manual_transaction_date(value),
    }


WEEKDAYS_SHORT = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


def _format_recurring_payment(payment) -> str:
    amount = float(payment[3] or 0)
    amount_prefix = "~" if payment[4] else ""
    repeat_type = payment[8]
    if repeat_type == "monthly":
        repeat = f"каждый месяц {payment[9] or 1} числа"
    elif repeat_type == "weekly":
        weekday = payment[10] if payment[10] is not None else 0
        repeat = f"каждую неделю, {WEEKDAYS_SHORT[int(weekday)]}"
    else:
        repeat = repeat_type or "без повтора"
    return f"{payment[2]} — {amount_prefix}{amount:,.0f} ₽, {repeat}"


def _next_recurring_date(repeat_type: str, day_month: int | None, day_week: int | None) -> date:
    today = date.today()
    if repeat_type == "monthly":
        day = max(1, min(28, int(day_month or 1)))
        year = today.year
        month = today.month
        if today.day >= day:
            month += 1
            if month > 12:
                month = 1
                year += 1
        return date(year, month, day)
    if repeat_type == "weekly":
        weekday = int(day_week if day_week is not None else today.weekday())
        days_ahead = (weekday - today.weekday()) % 7
        if days_ahead == 0:
            days_ahead = 7
        return today + timedelta(days=days_ahead)
    return today + timedelta(days=1)


async def render_categories_text(user_id: int) -> str:
    cats = await get_categories(user_id)
    if not cats:
        return "Категорий пока нет."

    income = [c for c in cats if c.get("type") == "income"]
    expense = [c for c in cats if c.get("type") == "expense"]

    text = "📂 Твои категории\n\n"
    text += "💰 Доходы:\n"
    text += "\n".join(["• " + c["name"] for c in income]) if income else "• пока нет"
    text += "\n\n💸 Расходы:\n"
    text += "\n".join(["• " + c["name"] for c in expense]) if expense else "• пока нет"
    text += (
        "\n\nЧтобы изменить статьи, открой «Ручной ввод» → «Расход» или «Доход» "
        "и напиши: «Заменить Кафе на Рестораны», «Добавить Подарки» или «Удалить Старую статью»."
        "\n\nМожно не разбираться в меню: открой ИИ-помощник и обычными словами попроси настроить категории, "
        "перенести транзакции, создать календарь платежей или поправить старые записи."
    )
    return text


async def _send_welcome_voice(message: Message):
    from app.database import get_bot_setting

    try:
        voice_file_id = await get_bot_setting("welcome_voice_file_id")
    except Exception:
        return
    if not voice_file_id:
        return

    try:
        await message.answer_voice(voice_file_id)
    except Exception:
        return


@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    user = message.from_user
    await get_or_create_user(
        user.id,
        user.username or "",
        user.full_name or "",
        user.language_code or "ru",
    )
    await _send_welcome_voice(message)

    from app.handlers.ai_assistant import AI_MODE_PROFILE, AIState

    intro = (
        "👋 <b>Я Баланс.</b>\n\n"
        "Деньги любят ясность. А ещё они почему-то любят исчезать, когда их не записывают.\n\n"
        "<b>Быстро знакомимся:</b>\n"
        "• какая у тебя финансовая цель;\n"
        "• что бесит или ломается в деньгах;\n"
        "• какие привычки, страхи или хаос надо приручить.\n\n"
        "Напиши одним сообщением. Честно можно: я не налоговая."
    )
    await state.set_state(AIState.chatting)
    await state.update_data(
        ai_mode=AI_MODE_PROFILE,
        history=[{"role": "assistant", "content": intro}],
    )
    await message.answer(
        intro,
        parse_mode="HTML",
        reply_markup=onboarding_finish_keyboard(),
    )


@router.message(F.text.in_({"🏠 Меню", "Меню", "/start"}))
async def msg_open_menu(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("Выбирай действие:", reply_markup=main_menu())


@router.callback_query(F.data == "main_menu")
async def cb_main_menu(call: CallbackQuery, state: FSMContext):
    await state.clear()
    try:
        await call.message.edit_text("Выбирай действие:", reply_markup=main_menu())
    except Exception:
        await call.message.answer("Выбирай действие:", reply_markup=main_menu())


@router.callback_query(F.data == "categories_list")
async def cb_categories_list(call: CallbackQuery):
    await call.message.edit_text(
        await render_categories_text(call.from_user.id),
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Быстрая смена статей", callback_data="quick_category_rename")],
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ]),
    )


@router.callback_query(F.data == "quick_category_rename")
async def cb_quick_category_rename(call: CallbackQuery, state: FSMContext):
    await state.clear()
    cats = await get_categories(call.from_user.id)
    if not cats:
        await call.message.edit_text(
            "Категорий пока нет.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    income = [cat for cat in cats if cat.get("type") == "income"]
    expense = [cat for cat in cats if cat.get("type") == "expense"]
    buttons = []
    if expense:
        buttons.append([InlineKeyboardButton(text="Расходы", callback_data="noop")])
        buttons.extend([
            [InlineKeyboardButton(text=cat["name"], callback_data=f"quick_category_pick:{cat['id']}")]
            for cat in expense
        ])
    if income:
        buttons.append([InlineKeyboardButton(text="Доходы", callback_data="noop")])
        buttons.extend([
            [InlineKeyboardButton(text=cat["name"], callback_data=f"quick_category_pick:{cat['id']}")]
            for cat in income
        ])
    buttons.append([InlineKeyboardButton(text="Назад", callback_data="categories_list")])
    buttons.append([InlineKeyboardButton(text="Меню", callback_data="main_menu")])

    await call.message.edit_text(
        "Выбери статью, которую хочешь переименовать.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons),
    )


@router.callback_query(F.data.startswith("quick_category_pick:"))
async def cb_quick_category_pick(call: CallbackQuery, state: FSMContext):
    cat_id = int(call.data.split(":")[1])
    row = await fetchone(
        "SELECT id, name, type FROM categories WHERE id=%s AND user_id=%s",
        (cat_id, call.from_user.id),
    )
    if not row:
        await call.answer("Категория не найдена.", show_alert=True)
        return

    category_id, category_name, category_type = row
    await state.set_state(QuickCategoryRenameState.waiting_new_name)
    await state.update_data(
        quick_category_id=category_id,
        quick_category_name=category_name,
        quick_category_type=category_type,
    )
    await call.message.edit_text(
        f"Переименовываем статью «{category_name}».\n\n"
        "Напиши новое название категории одним сообщением.\n"
        "Я сохраню его с большой буквы и без точки в конце.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Выбрать другую", callback_data="quick_category_rename")],
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ]),
    )


@router.message(QuickCategoryRenameState.waiting_new_name, F.text)
async def msg_quick_category_new_name(message: Message, state: FSMContext):
    from app.services.category_commands import normalize_category_display_name

    data = await state.get_data()
    category_id = data.get("quick_category_id")
    old_name = data.get("quick_category_name")
    category_type = data.get("quick_category_type")
    new_name = normalize_category_display_name(message.text or "")

    if not category_id or not old_name or not category_type:
        await state.clear()
        await message.answer(
            "Не нашёл выбранную статью. Начни быструю смену заново.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Быстрая смена статей", callback_data="quick_category_rename")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return
    if not new_name:
        await message.answer("Напиши новое название категории, например: «Развитие».")
        return

    duplicate = await fetchone(
        "SELECT id FROM categories WHERE user_id=%s AND type=%s AND lower(name)=lower(%s) AND id<>%s",
        (message.from_user.id, category_type, new_name, category_id),
    )
    if duplicate:
        await message.answer(
            f"Статья «{new_name}» уже есть. Напиши другое название.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Выбрать другую", callback_data="quick_category_rename")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    await execute(
        "UPDATE categories SET name=%s WHERE id=%s AND user_id=%s",
        (new_name, category_id, message.from_user.id),
    )
    await state.clear()
    await message.answer(
        f"Готово: «{old_name}» теперь «{new_name}».",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Быстрая смена статей", callback_data="quick_category_rename")],
            [InlineKeyboardButton(text="Показать категории", callback_data="categories_list")],
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ]),
    )


@router.callback_query(F.data == "manual_input")
async def cb_manual_input(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await call.message.edit_text(
        "✍️ <b>Ручной ввод</b>\n\n"
        "Сначала выбираешь тип и категорию кнопками, потом пишешь или говоришь сумму с комментарием.\n"
        "Это принцип «ввод без ошибок»: категория не угадывается и не попадает случайно в «нету».\n\n"
        "Пример после выбора категории: <i>500 р блабла ла коммент</i>",
        parse_mode="HTML",
        reply_markup=manual_input_keyboard(),
    )


@router.callback_query(F.data == "add_expense")
async def cb_add_expense(call: CallbackQuery, state: FSMContext):
    from app.database import get_user_tier
    tier = await get_user_tier(call.from_user.id)
    if tier == 'free':
        await call.answer(show_alert=True, text=(
            "Пробный период закончился! Оформи подписку."
        ))
        await call.message.edit_text(
            "Пробный период закончился.\n\nДля записи транзакций нужен тариф Старт или Премиум.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="⭐ Тарифы", callback_data="premium")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )
        return
    categories = await get_categories(call.from_user.id, type_="expense")
    await state.set_state(AddTransaction.choosing_category)
    await state.update_data(tx_type="expense")
    await call.message.edit_text(
        await render_categories_text(call.from_user.id) + "\n\nВыбери категорию расхода:",
        reply_markup=categories_keyboard(categories, "cat", back_callback="manual_input"),
    )


@router.callback_query(F.data == "add_income")
async def cb_add_income(call: CallbackQuery, state: FSMContext):
    from app.database import get_user_tier
    tier = await get_user_tier(call.from_user.id)
    if tier == 'free':
        await call.answer(show_alert=True, text=(
            "Пробный период закончился! Оформи подписку."
        ))
        await call.message.edit_text(
            "Пробный период закончился.\n\nДля записи транзакций нужен тариф Старт или Премиум.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="⭐ Тарифы", callback_data="premium")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )
        return
    categories = await get_categories(call.from_user.id, type_="income")
    await state.set_state(AddTransaction.choosing_category)
    await state.update_data(tx_type="income")
    await call.message.edit_text(
        await render_categories_text(call.from_user.id) + "\n\nВыбери источник дохода:",
        reply_markup=categories_keyboard(categories, "cat", back_callback="manual_input"),
    )


@router.message(AddTransaction.choosing_category, F.text)
async def msg_category_command(message: Message, state: FSMContext):
    from app.services.category_commands import parse_category_command, apply_category_command

    data = await state.get_data()
    scope_type = data.get("tx_type")
    command = await parse_category_command(message.from_user.id, message.text or "", scope_type)
    if not command or command.get("intent") == "unknown":
        await message.answer(
            "Не понял команду. Можно выбрать категорию кнопкой или написать: «замени категорию X на Y».",
        )
        return

    result = await apply_category_command(message.from_user.id, command, scope_type)
    categories = await get_categories(message.from_user.id, type_=scope_type)
    title = "Выбери категорию расхода:" if scope_type == "expense" else "Выбери источник дохода:"
    await message.answer(
        result + "\n\n" + title,
        reply_markup=categories_keyboard(categories, "cat", back_callback="manual_input"),
    )


@router.callback_query(F.data.startswith("cat:"))
async def cb_category_selected(call: CallbackQuery, state: FSMContext):
    _, cat_id, kind = call.data.split(":")
    row = await fetchone(
        "SELECT name, type, kind FROM categories WHERE id=%s AND user_id=%s",
        (int(cat_id), call.from_user.id),
    )
    if not row:
        await call.answer("Категория не найдена.", show_alert=True)
        return
    category_name, category_type, category_kind = row
    await state.update_data(
        category_id=int(cat_id),
        category_name=category_name,
        tx_type=category_type,
        kind=category_kind or kind,
    )
    await state.set_state(AddTransaction.entering_amount)
    await call.message.edit_text(
        "Ручной ввод включён.\n\n"
        f"Категория: «{category_name}»\n\n"
        "Теперь напиши или скажи голосом сумму и комментарий одним сообщением.\n"
        "Например: «500 р блабла ла коммент»."
    )


@router.message(AddTransaction.entering_amount, F.voice)
async def msg_safe_manual_voice(message: Message, state: FSMContext):
    from app.handlers.voice import format_recognized_text, transcribe_voice

    thinking = await message.answer("Распознаю голос...")
    try:
        file = await message.bot.get_file(message.voice.file_id)
        file_bytes = await message.bot.download_file(file.file_path)
        text = await transcribe_voice(file_bytes.read())
        if not text:
            await thinking.edit_text("Не удалось распознать голос. Попробуй ещё раз.")
            return
        await thinking.edit_text(format_recognized_text(text), parse_mode="HTML")
        await _save_safe_manual_transaction(message, state, text)
    except Exception as e:
        await thinking.edit_text("Ошибка распознавания: " + str(e))


@router.message(AddTransaction.entering_amount, F.text)
async def msg_amount(message: Message, state: FSMContext):
    await _save_safe_manual_transaction(message, state, message.text or "")


async def _save_safe_manual_transaction(message: Message, state: FSMContext, text: str):
    parsed = _parse_safe_manual_input(text)
    if not parsed:
        await message.answer(
            "Не нашёл сумму. Напиши, например: «500 р комментарий» или просто «500»."
        )
        return
    await state.update_data(
        amount=parsed["amount"],
        transaction_date=parsed["transaction_date"],
    )
    await _save_transaction(message, state, comment=parsed["comment"])


@router.message(AddTransaction.entering_comment, Command("skip"))
async def msg_skip_comment(message: Message, state: FSMContext):
    await _save_transaction(message, state, comment="")


@router.message(AddTransaction.entering_comment)
async def msg_comment(message: Message, state: FSMContext):
    await _save_transaction(message, state, comment=message.text or "")


async def _save_transaction(message: Message, state: FSMContext, comment: str):
    data = await state.get_data()
    tx = await create_transaction(
        user_id=message.from_user.id,
        category_id=data["category_id"],
        amount=data["amount"],
        type_=data["tx_type"],
        kind=data["kind"],
        comment=comment,
        transaction_date=data.get("transaction_date"),
    )
    await state.clear()
    sign = "−" if data["tx_type"] == "expense" else "+"
    tx_date = data.get("transaction_date") or date.today()
    category_name = data.get("category_name") or "выбранная категория"
    text = (
        f"✅ Записано!\n\n"
        f"{sign}{data['amount']:,.0f} ₽\n"
        f"📂 {category_name}\n"
        f"📅 {tx_date.strftime('%d.%m.%Y')}\n"
        f"{_format_comment_block(comment)}"
    )
    insight = await build_transaction_insight(message.from_user.id, tx["id"])
    if insight:
        text += "\n\n" + insight
    await message.answer(
        text,
        reply_markup=confirm_keyboard(tx["id"]),
    )


@router.callback_query(F.data == "report_month")
async def cb_report_month(call: CallbackQuery):
    now = datetime.now()
    MN = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
          7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    MONTHS_RU = {1:"Yanvar",2:"Fevral",3:"Mart",4:"Aprel",5:"Mai",6:"Iyun",
                 7:"Iyul",8:"Avgust",9:"Sentyabr",10:"Oktyabr",11:"Noyabr",12:"Dekabr"}
    months = []
    for i in range(3):
        m = now.month - i
        y = now.year
        if m <= 0:
            m += 12
            y -= 1
        months.append((y, m))
    kb = InlineKeyboardMarkup(inline_keyboard=[
        *[[InlineKeyboardButton(
            text=MONTHS_RU[m] + " " + str(y),
            callback_data="report:" + str(y) + ":" + str(m)
        )] for y, m in months],
        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
    ])
    await call.message.edit_text("Отчёт ДДС - выбери месяц:", parse_mode=None, reply_markup=kb)
    return


async def render_month_report(user_id: int, year: int, month: int):
    summary = await get_monthly_summary(user_id, year, month)
    breakdown = await get_category_breakdown(user_id, year, month)
    month_name = str(year) + "-" + str(month).zfill(2)
    actual_expense = summary["total_expense"]
    cash_balance = summary["income"] - actual_expense
    text = (
        f"📊 <b>Отчёт за {month_name}</b>\n\n"
        f"💰 Доходы:             <b>{summary['income']:,.0f} ₽</b>\n"
        f"🛒 Расходы:            <b>{actual_expense:,.0f} ₽</b>\n"
        f"━━━━━━━━━━━━━━━\n"
        f"{'✅' if cash_balance >= 0 else '🔴'} Остаток ДС: "
        f"<b>{cash_balance:+,.0f} ₽</b>\n\n"
    )
    # Категории с % от доходов (если доходов нет — % от фактических расходов)
    income = summary['income']
    total_expense = actual_expense
    pct_base = income if income > 0 else (total_expense if total_expense > 0 else 0)
    balance = cash_balance
    balance_pct = (balance / pct_base * 100) if pct_base else 0

    # Получаем все категории пользователя
    all_cats = await get_categories(user_id)
    expense_cats = {
        cat['name']: 0.0
        for cat in all_cats
        if cat.get('type') == 'expense'
    }

    # Заполняем суммами и kind из breakdown
    for row in breakdown:
        name = row['name']
        if name in expense_cats:
            expense_cats[name] = float(row['total'])

    # Сортируем по убыванию
    sorted_cats = sorted(expense_cats.items(), key=lambda x: x[1], reverse=True)

    if sorted_cats:
        text += "📋 <b>Расходы:</b>\n"
        for name, total in sorted_cats:
            pct = (total / pct_base * 100) if pct_base else 0
            text += f"  🛒 {name}: {total:,.0f} ₽ ({pct:.0f}%)\n"

    # Добавляем % к остатку
    text = text.replace(
        f"<b>{balance:+,.0f} ₽</b>",
        f"<b>{balance:+,.0f} ₽ ({balance_pct:.0f}%)</b>"
    )
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="📈 График", callback_data="chart:" + str(year) + ":" + str(month))],
        [InlineKeyboardButton(text="🤖 ИИ-помощник", callback_data="ai_assistant")],
        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
    ])
    return text, kb


@router.callback_query(F.data.startswith("report:"))
async def cb_report_by_month(call: CallbackQuery):
    parts = call.data.split(":")
    year, month = int(parts[1]), int(parts[2])
    text, kb = await render_month_report(call.from_user.id, year, month)
    await call.message.edit_text(text, parse_mode="HTML", reply_markup=kb)


@router.callback_query(F.data == "recent")
async def cb_recent(call: CallbackQuery):
    now = datetime.now()
    MONTHS = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
              7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
    months = []
    for i in range(3):
        m = now.month - i
        y = now.year
        if m <= 0:
            m += 12
            y -= 1
        months.append((y, m))
    kb = InlineKeyboardMarkup(inline_keyboard=[
        *[[InlineKeyboardButton(
            text=f"{MONTHS[m]} {y}",
            callback_data=f"txlist:{y}:{m}"
        )] for y, m in months],
        [InlineKeyboardButton(text="Выгрузить всё в Excel", callback_data="export_all")],
        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
    ])
    await call.message.edit_text(
        "📋 <b>Последние операции</b>\n\n"
        "Здесь история твоих расходов и доходов по месяцам.\n"
        "Операции можно смотреть, удалять и исправлять: сумму, дату, категорию или комментарий.\n\n"
        "Выбери месяц:",
        parse_mode="HTML",
        reply_markup=kb,
    )



@router.callback_query(F.data == "ai_analyze")
async def cb_ai_analyze(call: CallbackQuery):
    prem = await is_premium(call.from_user.id)
    if not prem:
        await call.answer("Доступно только в Premium ⭐", show_alert=True)
        return
    await call.answer("ИИ-помощник уже доступен в меню.", show_alert=True)


@router.callback_query(F.data == "scan_receipt")
async def cb_scan_receipt(call: CallbackQuery):
    prem = await is_premium(call.from_user.id)
    if not prem:
        await call.answer("Доступно только в Premium ⭐", show_alert=True)
        return
    await call.answer("Сканирование чеков — скоро!", show_alert=True)


@router.callback_query(F.data.startswith("delete_tx:"))
async def cb_delete_tx(call: CallbackQuery):
    tx_id = int(call.data.split(":")[1])
    public_id = await public_number_for_transaction(call.from_user.id, tx_id)
    await execute(
        "DELETE FROM transactions WHERE id=%s AND user_id=%s",
        (tx_id, call.from_user.id)
    )
    await call.message.edit_text(f"Транзакция #{public_id} удалена.", reply_markup=main_menu())


@router.callback_query(F.data.startswith("confirm_delete_intent:"))
async def cb_confirm_delete_intent(call: CallbackQuery):
    from app.database import delete_transaction_by_id

    tx_id = int(call.data.split(":")[1])
    public_id = await public_number_for_transaction(call.from_user.id, tx_id)
    success = await delete_transaction_by_id(call.from_user.id, tx_id)
    text = f"Транзакция #{public_id} удалена." if success else "Транзакция не найдена."
    await call.message.edit_text(text, reply_markup=main_menu())


@router.callback_query(F.data == "cancel_delete_intent")
async def cb_cancel_delete_intent(call: CallbackQuery, state: FSMContext):
    await state.set_state(DeleteTxState.waiting_id)
    await state.update_data(delete_year=None, delete_month=None)
    await call.message.edit_text(
        "Хорошо. Введите номер транзакции для удаления, например: 712.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ]),
    )


@router.callback_query(F.data.startswith("confirm_change_tx_category:"))
async def cb_confirm_change_tx_category(call: CallbackQuery):
    _, tx_id, category_id = call.data.split(":")
    row = await fetchone(
        "SELECT name, type, kind FROM categories WHERE id=%s AND user_id=%s",
        (int(category_id), call.from_user.id),
    )
    if not row:
        await call.message.edit_text("Категория не найдена.", reply_markup=main_menu())
        return
    category_name, type_, kind = row
    await execute(
        "UPDATE transactions SET category_id=%s, type=%s, kind=%s WHERE id=%s AND user_id=%s",
        (int(category_id), type_, kind, int(tx_id), call.from_user.id),
    )
    public_id = await public_number_for_transaction(call.from_user.id, int(tx_id))
    await call.message.edit_text(
        f"Готово. Транзакция #{public_id} перенесена в категорию «{category_name}».",
        reply_markup=main_menu(),
    )


@router.callback_query(F.data == "cancel_change_tx_category")
async def cb_cancel_change_tx_category(call: CallbackQuery):
    await call.message.edit_text(
        "Смена категории отменена. Можно написать номер транзакции и новую категорию точнее.",
        reply_markup=main_menu(),
    )


@router.callback_query(F.data.startswith("confirm:"))
async def cb_confirm(call: CallbackQuery):
    await call.message.edit_reply_markup(reply_markup=main_menu())


@router.callback_query(F.data == "noop")
async def cb_noop(call: CallbackQuery):
    await call.answer()


def _tx_row_to_current(tx) -> dict:
    if isinstance(tx, dict):
        return {
            "id": tx.get("id"),
            "amount": float(tx.get("amount") or 0),
            "type": tx.get("type") or "expense",
            "comment": tx.get("comment") or "",
            "transaction_date": tx.get("transaction_date"),
            "category_id": tx.get("category_id"),
            "category_name": tx.get("cat_name") or tx.get("category_name") or "",
        }
    return {
        "id": tx[0],
        "amount": float(tx[1] or 0),
        "type": tx[2] or "expense",
        "comment": tx[3] or "",
        "transaction_date": tx[4],
        "category_id": tx[5] if len(tx) > 5 else None,
        "category_name": tx[6] if len(tx) > 6 else "",
    }


def _format_current_tx(current: dict) -> str:
    sign = "-" if current["type"] == "expense" else "+"
    date_value = current["transaction_date"].strftime("%d.%m")
    comment = _format_comment_inline(current.get("comment"))
    return f"{date_value} {sign}{abs(float(current['amount'])):,.0f} {current.get('category_name') or ''}{comment}"


async def _build_transaction_edit(user_id: int, text: str, tx_id: int):
    from app.database import get_categories
    from app.services.transaction_commands import parse_transaction_edit

    tx = await fetchone(
        "SELECT t.id, t.amount, t.type, t.comment, t.transaction_date, t.category_id, c.name as cat_name "
        "FROM transactions t LEFT JOIN categories c ON t.category_id = c.id "
        "WHERE t.id = %s AND t.user_id = %s",
        (tx_id, user_id),
    )
    if not tx:
        return None, None

    current = _tx_row_to_current(tx)
    parsed = await parse_transaction_edit(user_id, text, current)

    amount = parsed.get("amount")
    category_id = parsed.get("new_category_id")
    comment = parsed.get("comment")
    date_value = current["transaction_date"]
    if parsed.get("day") and parsed.get("month"):
        try:
            date_value = datetime(
                int(parsed.get("year") or datetime.now().year),
                int(parsed["month"]),
                int(parsed["day"]),
            ).date()
        except ValueError:
            date_value = current["transaction_date"]

    if amount is None:
        amount = current["amount"]
    if comment is None:
        comment = current["comment"]
    if not category_id:
        category_id = current["category_id"]

    type_ = current["type"]
    category_name = current["category_name"]
    cats = await get_categories(user_id)
    for cat in cats:
        if cat["id"] == category_id:
            type_ = cat.get("type", type_)
            category_name = cat.get("name", category_name)
            break

    updated = {
        "tx_id": tx_id,
        "amount": float(amount),
        "type": type_,
        "category_id": category_id,
        "category_name": category_name,
        "comment": comment,
        "transaction_date": date_value,
    }
    if (
        updated["amount"] == current["amount"]
        and updated["type"] == current["type"]
        and updated["category_id"] == current["category_id"]
        and updated["comment"] == current["comment"]
        and updated["transaction_date"] == current["transaction_date"]
    ):
        return current, None
    return current, updated


async def _apply_transaction_edit(user_id: int, edit: dict):
    if not edit:
        return False
    await execute(
        "UPDATE transactions SET amount=%s, type=%s, category_id=%s, comment=%s, transaction_date=%s WHERE id=%s AND user_id=%s",
        (
            edit["amount"],
            edit["type"],
            edit["category_id"],
            edit["comment"],
            edit["transaction_date"],
            edit["tx_id"],
            user_id,
        ),
    )
    return True


@router.callback_query(F.data == "confirm_edit_transaction")
async def cb_confirm_edit_transaction(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    edit = data.get("pending_transaction_edit")
    success = await _apply_transaction_edit(call.from_user.id, edit)
    await state.update_data(pending_transaction_edit=None)
    await call.message.edit_text(
        "Готово, операция обновлена." if success else "Не нашёл правку для применения.",
        reply_markup=main_menu(),
    )


@router.callback_query(F.data == "cancel_edit_transaction")
async def cb_cancel_edit_transaction(call: CallbackQuery, state: FSMContext):
    await state.update_data(pending_transaction_edit=None)
    await call.message.edit_text("Ок, не меняю операцию.", reply_markup=main_menu())


@router.callback_query(F.data.startswith("confirm_delete_recurring_intent:"))
async def cb_confirm_delete_recurring_intent(call: CallbackQuery):
    payment_id = int(call.data.split(":")[1])
    await execute(
        "DELETE FROM recurring_payments WHERE id=%s AND user_id=%s",
        (payment_id, call.from_user.id),
    )
    await call.message.edit_text("Платёж удалён из календаря.", reply_markup=main_menu())


@router.callback_query(F.data == "confirm_edit_recurring_intent")
async def cb_confirm_edit_recurring_intent(call: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    edit = data.get("pending_recurring_edit")
    if not edit:
        await call.message.edit_text("Не нашёл правку для применения.", reply_markup=main_menu())
        return

    await execute(
        """
        UPDATE recurring_payments
        SET amount=%s,
            amount_is_approximate=%s,
            repeat_type=%s,
            repeat_day_of_month=%s,
            repeat_day_of_week=%s,
            next_trigger_date=%s
        WHERE id=%s AND user_id=%s
        """,
        (
            edit["amount"],
            edit["amount_is_approximate"],
            edit["repeat_type"],
            edit["repeat_day_of_month"],
            edit["repeat_day_of_week"],
            edit["next_trigger_date"],
            edit["payment_id"],
            call.from_user.id,
        ),
    )
    await state.update_data(pending_recurring_edit=None)
    await call.message.edit_text("Готово, платёжный календарь обновлён.", reply_markup=main_menu())


@router.callback_query(F.data == "cancel_edit_recurring_intent")
async def cb_cancel_edit_recurring_intent(call: CallbackQuery, state: FSMContext):
    await state.update_data(pending_recurring_edit=None)
    await call.message.edit_text("Ок, календарь не меняю.", reply_markup=main_menu())


@router.callback_query(F.data == "settings")
async def cb_settings(call: CallbackQuery):
    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="◀️ Назад", callback_data="main_menu")]
    ])
    await call.message.edit_text(
        "⚙️ <b>Настройки</b>\n\nУправление категориями — скоро!",
        parse_mode="HTML",
        reply_markup=kb,
    )


async def _send_beautiful_report_result(bot, chat_id: int, task_id: str) -> bool:
    from app.services.kie_reports import extract_result_url, get_kie_task

    try:
        task = await get_kie_task(task_id)
    except Exception:
        return False

    state = task.get("state") or "unknown"
    if state == "success":
        image_url = extract_result_url(task)
        if image_url:
            await bot.send_photo(
                chat_id,
                image_url,
                caption="Готово! Красивый отчёт собран.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
        else:
            await bot.send_message(
                chat_id,
                "Отчёт собрался, но я не смог получить картинку. Попробуй запросить отчёт ещё раз чуть позже.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
        return True

    if state == "fail":
        await bot.send_message(
            chat_id,
            "Не получилось собрать красивый отчёт. Попробуй ещё раз чуть позже.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return True

    return False


async def _wait_and_send_beautiful_report(bot, chat_id: int, task_id: str):
    for _ in range(36):
        await asyncio.sleep(10)
        if await _send_beautiful_report_result(bot, chat_id, task_id):
            return

    await bot.send_message(
        chat_id,
        "Отчёт всё ещё готовится. Я пока не получил картинку, но можно запросить красивый отчёт ещё раз чуть позже.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ]),
    )


@router.callback_query(F.data.startswith("kie_report:"))
async def cb_kie_report_ready(call: CallbackQuery):
    task_id = call.data.split(":", 1)[1]
    await call.answer("Проверяю отчёт...")
    if not await _send_beautiful_report_result(call.bot, call.message.chat.id, task_id):
        await call.message.answer(
            "Отчёт ещё готовится. Я пришлю картинку сюда, когда она будет готова.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        asyncio.create_task(_wait_and_send_beautiful_report(call.bot, call.message.chat.id, task_id))


def _beautiful_report_confirm_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Сформировать", callback_data="beautiful_report_generate")],
        [InlineKeyboardButton(text="Отмена", callback_data="beautiful_report_cancel")],
    ])


async def _start_beautiful_report_generation(
    message: Message,
    state: FSMContext,
    user_id: int,
    year: int,
    month: int,
    user_prompt: str,
):
    from app.database import get_user_tier
    from app.services.kie_reports import start_beautiful_report

    tier = await get_user_tier(user_id)
    if tier == "free":
        await state.clear()
        await message.answer(
            "Красивые отчёты доступны на платных тарифах.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Тарифы", callback_data="premium")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    thinking = await message.answer("Собираю данные и запускаю красивый отчёт...")
    try:
        result = await start_beautiful_report(user_id, year, month, user_prompt)
        await state.clear()
        await thinking.edit_text(
            "Принял. Собираю красивый отчёт и пришлю картинку сюда, когда она будет готова.\n\n"
            "Обычно это занимает около минуты.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Обычный отчёт", callback_data="report:" + str(year) + ":" + str(month))],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        asyncio.create_task(_wait_and_send_beautiful_report(message.bot, message.chat.id, result["task_id"]))
    except Exception:
        await state.clear()
        await thinking.edit_text(
            "Не смог запустить красивый отчёт. Проверь настройки генерации отчётов или попробуй позже.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Обычный отчёт", callback_data="report:" + str(year) + ":" + str(month))],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )


@router.callback_query(F.data == "beautiful_report_generate")
async def cb_beautiful_report_generate(call: CallbackQuery, state: FSMContext):
    await call.answer()
    data = await state.get_data()
    report_data = data.get("beautiful_report") or {}
    year = report_data.get("year")
    month = report_data.get("month")
    user_prompt = report_data.get("user_prompt") or "Сделай красивый финансовый отчёт."
    if not year or not month:
        await state.clear()
        await call.message.answer(
            "Не нашёл параметры отчёта. Запроси красивый отчёт ещё раз.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return
    await _start_beautiful_report_generation(
        call.message,
        state,
        call.from_user.id,
        int(year),
        int(month),
        user_prompt,
    )


@router.callback_query(F.data == "beautiful_report_cancel")
async def cb_beautiful_report_cancel(call: CallbackQuery, state: FSMContext):
    await call.answer("Отменил")
    await state.clear()
    await call.message.edit_text(
        "Ок, красивый отчёт не запускаю.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ]),
    )


@router.message(BeautifulReportState.waiting_confirm, F.text)
async def msg_beautiful_report_refine(message: Message, state: FSMContext):
    from app.services.kie_reports import describe_report_plan

    data = await state.get_data()
    report_data = data.get("beautiful_report") or {}
    year = int(report_data.get("year") or datetime.now().year)
    month = int(report_data.get("month") or datetime.now().month)
    previous_prompt = report_data.get("user_prompt") or "Сделай красивый финансовый отчёт."
    lower = (message.text or "").lower().strip()
    if lower in ("отмена", "отмени", "назад", "завершить", "стоп"):
        await state.clear()
        await message.answer(
            "Ок, красивый отчёт не запускаю.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return
    if lower in ("сформировать", "формируй", "запускай", "да", "ок"):
        await _start_beautiful_report_generation(
            message,
            state,
            message.from_user.id,
            year,
            month,
            previous_prompt,
        )
        return
    updated_prompt = previous_prompt + "\nУточнение: " + (message.text or "").strip()
    await state.update_data(beautiful_report={
        "year": year,
        "month": month,
        "user_prompt": updated_prompt,
    })
    await message.answer(
        describe_report_plan(year, month, updated_prompt),
        reply_markup=_beautiful_report_confirm_keyboard(),
    )


# --- Быстрый ввод ---

@router.message(F.text.regexp(r'^[+-]?\d+(?![.\d])'), StateFilter(default_state))
async def msg_quick_input(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state:
        return

    from app.services.transaction_ai import extract_transactions_from_text

    transactions = await extract_transactions_from_text(message.from_user.id, message.text or "", source="text")
    if not transactions:
        return

    saved_items = []
    saved_ids = []
    for tx in transactions:
        saved = await create_transaction(
            user_id=message.from_user.id,
            category_id=tx["category_id"],
            amount=tx["amount"],
            type_=tx["type"],
            kind=tx.get("kind"),
            comment=tx.get("comment") or "",
            transaction_date=tx.get("transaction_date"),
            pnl_period=tx.get("pnl_period"),
        )
        saved_items.append((saved, tx))
        saved_ids.append(saved["id"])

    if len(saved_items) > 1:
        public_ids = await public_numbers_for_ids(message.from_user.id, saved_ids)
        lines = []
        for saved, tx in saved_items:
            sign = "−" if tx["type"] == "expense" else "+"
            public_id = public_ids.get(int(saved["id"]), str(saved["id"]))
            line = (
                f"{sign}{float(tx['amount']):,.0f} ₽ — "
                f"{escape(str(tx['category_name']))} #{escape(str(public_id))}"
            )
            if tx.get("comment"):
                line += " | 💬 «" + escape(str(tx["comment"])) + "»"
            lines.append(line)

        text = f"✅ <b>Записано {len(saved_items)} операций:</b>\n\n" + "\n".join(lines)
        insight = await build_first_transaction_insight(message.from_user.id, saved_ids)
        if insight:
            text += "\n\n" + insight
        await message.answer(
            text,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="📋 Открыть список", callback_data="recent")],
                [InlineKeyboardButton(text="🏠 Меню", callback_data="main_menu")],
            ]),
        )
        return

    saved, tx = saved_items[0]
    public_id = await public_number_for_transaction(message.from_user.id, saved["id"])

    sign = "−" if tx["type"] == "expense" else "+"
    date_str = tx["transaction_date"].strftime("%d.%m")

    text = (
        f"✅ <b>Записано!</b>\n\n"
        f"{sign}{tx['amount']:,.0f} ₽\n"
        f"📂 {escape(str(tx['category_name']))}\n"
        f"📅 {date_str}\n"
        f"🔢 #{escape(str(public_id))}\n"
    )
    if tx.get("comment"):
        text += _format_comment_block_html(tx["comment"]) + "\n"
    if tx.get("pnl_period"):
        text += f"📊 ПнЛ: {escape(str(tx['pnl_period']))}\n"

    insight = await build_transaction_insight(message.from_user.id, saved["id"])
    if insight:
        text += "\n" + insight + "\n"

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    await message.answer(
        text,
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Верно", callback_data=f"confirm:{saved['id']}"),
                InlineKeyboardButton(text="🗑 Удалить", callback_data=f"delete_tx:{saved['id']}"),
            ],
            [InlineKeyboardButton(text="🏠 Меню", callback_data="main_menu")],
        ])
    )


# --- /reset и /deleteaccount ---

@router.message(Command("reset"))
async def cmd_reset(message: Message):
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    await message.answer(
        "⚠️ <b>Сброс данных</b>\n\n"
        "Это удалит все твои транзакции и регулярные платежи.\n"
        "Категории и настройки сохранятся.\n\n"
        "Ты уверен?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Да, удалить данные", callback_data="confirm_reset"),
                InlineKeyboardButton(text="❌ Отмена", callback_data="main_menu"),
            ]
        ])
    )


@router.message(Command("deleteaccount"))
async def cmd_delete_account(message: Message):
    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    await message.answer(
        "🗑 <b>Удаление аккаунта</b>\n\n"
        "Это удалит ВСЕ твои данные без возможности восстановления:\n"
        "транзакции, категории, настройки, подписку.\n\n"
        "При следующем /start ты начнёшь заново.\n\n"
        "Ты уверен?",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Да, удалить аккаунт", callback_data="confirm_delete_account"),
                InlineKeyboardButton(text="❌ Отмена", callback_data="main_menu"),
            ]
        ])
    )


@router.callback_query(F.data == "confirm_reset")
async def cb_confirm_reset(call: CallbackQuery):
    await execute(
        "DELETE FROM transactions WHERE user_id=%s",
        (call.from_user.id,)
    )
    await execute(
        "DELETE FROM recurring_payments WHERE user_id=%s",
        (call.from_user.id,)
    )
    await call.message.edit_text(
        "✅ Все транзакции и регулярные платежи удалены.\n\n"
        "Можешь начинать заново!",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🏠 Меню", callback_data="main_menu")]
        ])
    )


@router.callback_query(F.data == "confirm_delete_account")
async def cb_confirm_delete_account(call: CallbackQuery):
    await execute(
        "DELETE FROM users WHERE id=%s",
        (call.from_user.id,)
    )
    await call.message.edit_text(
        "🗑 Аккаунт удалён. Все данные стёрты.\n\n"
        "Нажми /start чтобы начать заново."
    )


# --- Дашборд управленца ---

@router.callback_query(F.data == "dashboard")
async def cb_dashboard(call: CallbackQuery):
    from app.database import get_dashboard, can_use_feature
    if not await can_use_feature(call.from_user.id, 'business_tools'):
        await call.message.edit_text(
            "Табло управленца доступно на тарифе Business.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Тарифы", callback_data="premium")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )
        return

    now = datetime.now()
    d = await get_dashboard(call.from_user.id, now.year, now.month)

    MONTHS = {1:'Январь',2:'Февраль',3:'Март',4:'Апрель',5:'Май',6:'Июнь',
              7:'Июль',8:'Август',9:'Сентябрь',10:'Октябрь',11:'Ноябрь',12:'Декабрь'}

    # Динамика
    if d['dynamics'] is not None:
        dyn_icon = "📈" if d['dynamics'] >= 0 else "📉"
        dyn_str = f"{dyn_icon} {d['dynamics']:+.1f}% vs прошлый месяц"
    else:
        dyn_str = "Нет данных за прошлый месяц"

    # Ближайшие платежи
    upcoming_str = ""
    if d['upcoming']:
        upcoming_str = "\n\nБлижайшие платежи (7 дней):\n"
        for p in d['upcoming']:
            upcoming_str += f"  {p[2].strftime('%d.%m')} {p[0]} — {float(p[1]):,.0f} руб.\n"

    # Топ расходов
    top_cats = sorted([c for c in d['categories'] if c[1] in ('fixed','variable')],
                      key=lambda x: x[2], reverse=True)[:3]
    top_str = ""
    if top_cats:
        top_str = "\nТоп расходов:\n"
        for name, kind, total in top_cats:
            pct = total / d['income'] * 100 if d['income'] > 0 else 0
            top_str += f"  {name}: {total:,.0f} ({pct:.1f}%)\n"

    text = (
        f"Табло управленца — {MONTHS[now.month]} {now.year}\n\n"
        f"Выручка: {d['income']:,.0f} руб.\n"
        f"Расходы: -{d['total_expense']:,.0f} руб.\n\n"
        f"EBITDA: {d['ebitda']:,.0f} руб."
    )

    if d['depreciation'] > 0:
        text += f"\n  Амортизация: -{d['depreciation']:,.0f}"
    if d['tax'] > 0:
        text += f"\n  Налоги: -{d['tax']:,.0f}"
    if d['loan_body'] > 0 or d['loan_pct'] > 0:
        text += f"\n  Кредиты: -{d['loan_body']+d['loan_pct']:,.0f}"

    net_icon = "✅" if d['net_profit'] >= 0 else "🔴"
    text += (
        f"\n\n{net_icon} Чистая прибыль: {d['net_profit']:,.0f} руб. ({d['net_profit_pct']:.1f}%)\n"
        f"{dyn_str}\n"
        f"\nТранзакций: {d['tx_count']}"
        f"{top_str}"
        f"{upcoming_str}"
    )

    await call.message.edit_text(
        text,
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ])
    )


# --- Транзакции по месяцу ---

@router.callback_query(F.data.startswith("txlist:"))
async def cb_txlist(call: CallbackQuery):
    from app.database import get_transactions_by_month
    await call.answer()
    parts = call.data.split(":")
    _, year, month = parts[:3]
    year, month = int(year), int(month)
    offset = int(parts[3]) if len(parts) > 3 else 0
    txs = await get_transactions_by_month(call.from_user.id, year, month)

    MONTHS = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
              7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}

    if not txs:
        await call.message.edit_text(
            f"За {MONTHS[month]} {year} транзакций нет.",
            parse_mode=None,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Назад", callback_data="recent")]
            ])
        )
        return

    page_size = 30
    page = txs[offset:offset + page_size]
    public_ids = await public_numbers_for_ids(call.from_user.id, [tx[0] for tx in page])
    text = f"Транзакции за {MONTHS[month]} {year}\n"
    text += f"{offset + 1}-{min(offset + page_size, len(txs))} из {len(txs)}\n\n"
    for tx in page:
        tx_id, date, amount, type_, comment, cat_name, wallet = tx
        public_id = public_ids.get(int(tx_id), str(tx_id))
        sign = "-" if type_ == "expense" else "+"
        comment_str = _format_comment_inline(comment)
        line = f"#{public_id} {date.strftime('%d.%m')} {sign}{abs(float(amount)):,.0f} {cat_name or ''}{comment_str}\n"
        if len(text) + len(line) > 3500:
            break
        text += line

    buttons = []
    nav = []
    if offset > 0:
        nav.append(InlineKeyboardButton(text="◀ Назад", callback_data=f"txlist:{year}:{month}:{max(0, offset-page_size)}"))
    if offset + page_size < len(txs):
        nav.append(InlineKeyboardButton(text="Ещё ▶", callback_data=f"txlist:{year}:{month}:{offset+page_size}"))
    if nav:
        buttons.append(nav)
    buttons.extend([
        [InlineKeyboardButton(text="✏️ Изменить транзакцию", callback_data=f"edit_by_id:{year}:{month}")],
        [InlineKeyboardButton(text="🗑 Удалить транзакцию", callback_data=f"delete_by_id:{year}:{month}")],
        [InlineKeyboardButton(text="Назад", callback_data="recent")],
    ])

    await call.message.edit_text(
        text,
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )


# --- Редактирование транзакции ---

class EditTxState(StatesGroup):
    waiting_value = State()


@router.callback_query(F.data.startswith("edit_by_id:"))
async def cb_edit_by_id(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    year, month = int(parts[1]), int(parts[2])
    offset = int(parts[3]) if len(parts) > 3 else 0
    await show_edit_list(call, year, month, offset)


async def show_edit_list(call, year, month, offset=0):
    from app.database import get_transactions_by_month
    txs = await get_transactions_by_month(call.from_user.id, year, month)
    if not txs:
        await call.answer("Нет транзакций за этот месяц.")
        return

    page = txs[offset:offset+20]
    public_ids = await public_numbers_for_ids(call.from_user.id, [tx[0] for tx in page])
    buttons = []
    for tx in page:
        tx_id, date, amount, type_, comment, cat_name, wallet = tx
        public_id = public_ids.get(int(tx_id), str(tx_id))
        sign = "-" if type_ == "expense" else "+"
        label = f"#{public_id} {date.strftime('%d.%m')} {sign}{int(float(amount))} {cat_name or ''}"
        buttons.append([InlineKeyboardButton(text=label, callback_data=f"edit_pick:{tx_id}:{year}:{month}")])

    nav = []
    if offset > 0:
        nav.append(InlineKeyboardButton(text="◀ Назад", callback_data=f"edit_by_id:{year}:{month}:{offset-20}"))
    if offset + 20 < len(txs):
        nav.append(InlineKeyboardButton(text="Ещё ▶", callback_data=f"edit_by_id:{year}:{month}:{offset+20}"))
    if nav:
        buttons.append(nav)

    buttons.append([InlineKeyboardButton(text="Отмена", callback_data=f"txlist:{year}:{month}")])
    await call.message.edit_text(
        f"Выбери транзакцию ({offset+1}-{min(offset+20, len(txs))} из {len(txs)}):",
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )


@router.callback_query(F.data.startswith("edit_pick:"))
async def cb_edit_pick(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    tx_id, year, month = int(parts[1]), parts[2], parts[3]

    tx = await fetchone(
        "SELECT t.id, t.amount, t.type, t.comment, t.transaction_date, c.name as cat_name "
        "FROM transactions t LEFT JOIN categories c ON t.category_id = c.id "
        "WHERE t.id = %s AND t.user_id = %s",
        (tx_id, call.from_user.id)
    )
    if not tx:
        await call.answer("Транзакция не найдена.")
        return

    await state.update_data(tx_id=tx_id, year=year, month=month)
    await state.set_state(EditTxState.waiting_value)

    # fetchone возвращает dict или tuple — обрабатываем оба
    if isinstance(tx, dict):
        tx_type = tx.get('type', 'expense')
        tx_amount = tx.get('amount', 0)
        tx_cat = tx.get('cat_name', '')
        tx_comment = tx.get('comment', '')
        tx_date = tx.get('transaction_date', '')
    else:
        tx_type = tx[2]
        tx_amount = tx[1]
        tx_cat = tx[5] if len(tx) > 5 else ''
        tx_comment = tx[3] if len(tx) > 3 else ''
        tx_date = tx[4] if len(tx) > 4 else ''

    sign = "-" if tx_type == 'expense' else "+"
    current = (sign + str(int(float(tx_amount))) + " " +
               str(tx_cat or '') + " " +
               str(tx_comment or ''))

    await call.message.edit_text(
        "Текущая транзакция: " + current.strip() + chr(10) + chr(10) +
        "Напиши, что изменить, обычными словами:" + chr(10) +
        "«поставь категорию Развитие»" + chr(10) +
        "«исправь сумму на 1200 руб»" + chr(10) +
        "«измени комментарий на аптека»" + chr(10) + chr(10) +
        "Можно и старым форматом: 13.06 -500 Еда кофе с молоком",
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Отмена", callback_data=f"txlist:{year}:{month}")]
        ])
    )


async def _apply_edit_tx_text(message: Message, state: FSMContext, text: str):
    from app.database import get_categories
    from app.parser import parse_quick_input
    from app.services.transaction_commands import parse_transaction_edit
    data = await state.get_data()
    tx_id = data['tx_id']
    year = data.get('year', '')
    month = data.get('month', '')

    tx = await fetchone(
        "SELECT t.id, t.amount, t.type, t.comment, t.transaction_date, t.category_id, c.name as cat_name "
        "FROM transactions t LEFT JOIN categories c ON t.category_id = c.id "
        "WHERE t.id = %s AND t.user_id = %s",
        (tx_id, message.from_user.id)
    )
    if not tx:
        await state.clear()
        await message.answer("Транзакция не найдена.", reply_markup=main_menu())
        return

    if isinstance(tx, dict):
        current = {
            "id": tx.get("id"),
            "amount": float(tx.get("amount") or 0),
            "type": tx.get("type") or "expense",
            "comment": tx.get("comment") or "",
            "transaction_date": tx.get("transaction_date"),
            "category_id": tx.get("category_id"),
            "category_name": tx.get("cat_name") or "",
        }
    else:
        current = {
            "id": tx[0],
            "amount": float(tx[1] or 0),
            "type": tx[2] or "expense",
            "comment": tx[3] or "",
            "transaction_date": tx[4],
            "category_id": tx[5],
            "category_name": tx[6] if len(tx) > 6 else "",
        }

    ai_parsed = await parse_transaction_edit(message.from_user.id, text, current)
    amount = ai_parsed.get("amount")
    category_id = ai_parsed.get("new_category_id")
    matched_cat = None
    comment = ai_parsed.get("comment")
    date = current["transaction_date"]

    if ai_parsed.get("day") and ai_parsed.get("month"):
        try:
            date = datetime(
                int(ai_parsed.get("year") or datetime.now().year),
                int(ai_parsed["month"]),
                int(ai_parsed["day"]),
            ).date()
        except ValueError:
            date = current["transaction_date"]

    if category_id:
        matched_cat = {
            "id": category_id,
            "name": ai_parsed.get("new_category_name"),
        }

    if amount is None and not category_id and comment is None and date == current["transaction_date"]:
        parsed = parse_quick_input(text)
        if not parsed or not parsed.get('amount'):
            await message.answer(
                "Не удалось распознать. Напиши проще, например: «поставь категорию Развитие» "
                "или «исправь сумму на 1200 руб».",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Отмена", callback_data=f"txlist:{year}:{month}")]
                ]),
            )
            return
        amount = parsed['amount']
        date = parsed.get('transaction_date')
        hint = parsed.get('category_hint', '')
        comment = parsed.get('comment', '')
        type_ = parsed['type']
    else:
        hint = ""
        type_ = current["type"]
        if amount is None:
            amount = current["amount"]
        if comment is None:
            comment = current["comment"]

    cats = await get_categories(message.from_user.id)
    if not category_id:
        for cat in cats:
            if hint and (hint.lower().strip() in cat['name'].lower() or cat['name'].lower() in hint.lower().strip()):
                category_id = cat['id']
                matched_cat = cat
                break
    if not category_id:
        category_id = current["category_id"]

    if matched_cat:
        type_ = matched_cat.get('type', type_)
    elif category_id:
        for cat in cats:
            if cat["id"] == category_id:
                type_ = cat.get("type", type_)
                break

    try:
        await execute(
            "UPDATE transactions SET amount=%s, type=%s, category_id=%s, comment=%s, transaction_date=%s WHERE id=%s AND user_id=%s",
            (amount, type_, category_id, comment, date, tx_id, message.from_user.id)
        )
        await state.clear()
        await message.answer(
            "Транзакция обновлена.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Назад к транзакциям", callback_data=f"txlist:{year}:{month}")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )
    except Exception as e:
        await state.clear()
        await message.answer("Ошибка: " + str(e))


@router.message(EditTxState.waiting_value, F.text)
async def msg_edit_tx_value(message: Message, state: FSMContext):
    await _apply_edit_tx_text(message, state, message.text or "")


@router.message(EditTxState.waiting_value, F.voice)
async def msg_edit_tx_voice(message: Message, state: FSMContext):
    from app.handlers.voice import format_recognized_text, transcribe_voice

    thinking = await message.answer("Распознаю голос...")
    try:
        file = await message.bot.get_file(message.voice.file_id)
        file_bytes = await message.bot.download_file(file.file_path)
        text = await transcribe_voice(file_bytes.read())
        await thinking.delete()
    except Exception as e:
        await thinking.delete()
        await message.answer("Не удалось распознать голос: " + str(e))
        return

    if not text:
        await message.answer("Не удалось распознать голос. Попробуй ещё раз.")
        return

    await message.answer(format_recognized_text(text), parse_mode="HTML")
    await _apply_edit_tx_text(message, state, text)


# --- Удаление по номеру ---

class DeleteTxState(StatesGroup):
    waiting_id = State()


@router.callback_query(F.data.startswith("delete_by_id"))
async def cb_delete_by_id(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    year = int(parts[1]) if len(parts) > 2 else None
    month = int(parts[2]) if len(parts) > 2 else None
    await state.set_state(DeleteTxState.waiting_id)
    await state.update_data(delete_year=year, delete_month=month)
    example = "712"
    hint = ""
    if month:
        example = f"{month}12"
        hint = "\nМожно ввести полный номер из списка или просто порядковый номер внутри месяца."
    await call.message.answer(
        f"Напишите номер транзакции для удаления (например: {example}).{hint}",
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Отмена", callback_data="recent")]
        ])
    )


@router.message(DeleteTxState.waiting_id, F.text)
async def msg_delete_tx_by_id(message: Message, state: FSMContext):
    from app.database import delete_transaction_by_id
    raw_number = (message.text or "").strip().replace("#", "")
    try:
        int(raw_number)
    except ValueError:
        handled = await handle_intent_message(message, state, message.text, source="text")
        if not handled:
            await message.answer("Введи числовой номер транзакции, например: 712")
        return

    data = await state.get_data()
    tx_id = await resolve_transaction_reference(
        message.from_user.id,
        raw_number,
        year=data.get("delete_year"),
        month=data.get("delete_month"),
    )
    await state.clear()
    if not tx_id:
        await message.answer(
            f"Транзакция #{raw_number} не найдена или не принадлежит тебе.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Попробовать снова", callback_data="delete_by_id")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )
        return

    public_id = await public_number_for_transaction(message.from_user.id, tx_id)
    success = await delete_transaction_by_id(message.from_user.id, tx_id)
    if success:
        await message.answer(
            f"Транзакция #{public_id} удалена.",
            reply_markup=main_menu()
        )
    else:
        await message.answer(
            f"Транзакция #{raw_number} не найдена или не принадлежит тебе.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Попробовать снова", callback_data="delete_by_id")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )


@router.message(DeleteTxState.waiting_id, F.voice)
async def msg_delete_tx_by_voice(message: Message, state: FSMContext):
    from app.handlers.voice import format_recognized_text, transcribe_voice

    thinking = await message.answer("Распознаю голос...")
    try:
        file = await message.bot.get_file(message.voice.file_id)
        file_bytes = await message.bot.download_file(file.file_path)
        text = await transcribe_voice(file_bytes.read())
        await thinking.delete()
    except Exception as e:
        await thinking.delete()
        await message.answer("Не удалось распознать голос: " + str(e))
        return

    if not text:
        await message.answer("Не удалось распознать голос. Попробуй ещё раз.")
        return

    await message.answer(format_recognized_text(text), parse_mode="HTML")
    if text.strip().replace("#", "").isdigit():
        data = await state.get_data()
        from app.database import delete_transaction_by_id
        raw_number = text.strip().replace("#", "")
        tx_id = await resolve_transaction_reference(
            message.from_user.id,
            raw_number,
            year=data.get("delete_year"),
            month=data.get("delete_month"),
        )
        await state.clear()
        if not tx_id:
            await message.answer(
                f"Транзакция #{raw_number} не найдена или не принадлежит тебе.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Попробовать снова", callback_data="delete_by_id")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return
        public_id = await public_number_for_transaction(message.from_user.id, tx_id)
        success = await delete_transaction_by_id(message.from_user.id, tx_id)
        await message.answer(
            f"Транзакция #{public_id} удалена." if success else f"Транзакция #{raw_number} не найдена или не принадлежит тебе.",
            reply_markup=main_menu() if success else InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Попробовать снова", callback_data="delete_by_id")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    handled = await handle_intent_message(message, state, text, source="voice")
    if not handled:
        await message.answer("Не понял, какую транзакцию удалить. Можно сказать номер или фразу вроде «удали 45 рублей за проезд».")


# --- Экспорт в Excel ---

@router.callback_query(F.data == "export_excel")
async def cb_export_excel(call: CallbackQuery):
    from app.database import get_all_transactions_for_export
    import io
    try:
        import openpyxl
    except ImportError:
        await call.answer("openpyxl не установлен", show_alert=True)
        return

    txs = await get_all_transactions_for_export(call.from_user.id)
    if not txs:
        await call.answer("Транзакций нет", show_alert=True)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    ws.append(["#", "Дата", "Сумма", "Тип", "Категория", "Комментарий", "ПнЛ период"])

    for tx in txs:
        tx_id, date, amount, type_, cat, comment, pnl = tx
        ws.append([
            tx_id,
            date.strftime("%d.%m.%Y"),
            float(amount) if type_ == "income" else -float(amount),
            "Доход" if type_ == "income" else "Расход",
            cat or "",
            comment or "",
            pnl or "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from aiogram.types import BufferedInputFile
    await call.message.answer_document(
        BufferedInputFile(buf.read(), filename="transactions.xlsx"),
        caption="Все транзакции"
    )
    await call.answer()


# --- ПнЛ отчёт ---

@router.callback_query(F.data == "pnl_report")
async def cb_pnl_report(call: CallbackQuery):
    from app.database import get_pnl_report, can_use_feature
    if not await can_use_feature(call.from_user.id, 'pnl_table'):
        await call.message.edit_text(
            "ПнЛ отчёт доступен на тарифе Premium и выше.",
            parse_mode=None,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Тарифы", callback_data="premium")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )
        return

    now = datetime.now()
    d = await get_pnl_report(call.from_user.id, now.year, now.month)

    MONTHS = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
              7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}

    pct = d['pct']

    text = f"ПнЛ — {MONTHS[now.month]} {now.year}\n\n"

    # Выручка
    text += f"ВЫРУЧКА: {d['income']:,.0f} руб.\n"
    for name, total in d['income_cats']:
        text += f"  {name}: {total:,.0f} ({pct(total)})\n"

    expense_total = d['variable'] + d['fixed']
    expense_cats = sorted(d['variable_cats'] + d['fixed_cats'], key=lambda item: item[1], reverse=True)
    text += f"\nРАСХОДЫ: -{expense_total:,.0f} ({pct(expense_total)})\n"
    for name, total in expense_cats:
        text += f"  {name}: -{total:,.0f} ({pct(total)})\n"

    # EBITDA
    eb_icon = "+" if d['ebitda'] >= 0 else ""
    text += f"\nEBITDA: {eb_icon}{d['ebitda']:,.0f} ({pct(d['ebitda'])})\n"

    # Ниже EBITDA
    if d['depreciation'] > 0:
        text += f"  Амортизация: -{d['depreciation']:,.0f}\n"
    if d['tax'] > 0:
        text += f"  Налоги: -{d['tax']:,.0f}\n"
    if d['loan_body'] > 0:
        text += f"  Кредит (тело): -{d['loan_body']:,.0f}\n"
    if d['loan_pct'] > 0:
        text += f"  Кредит (проценты): -{d['loan_pct']:,.0f}\n"

    # ЧП
    np_icon = "+" if d['net_profit'] >= 0 else ""
    text += f"\nЧИСТАЯ ПРИБЫЛЬ: {np_icon}{d['net_profit']:,.0f} ({pct(d['net_profit'])})\n"

    await call.message.edit_text(
        text,
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ])
    )


# --- График ---

@router.callback_query(F.data == "chart_month")
async def cb_chart_month(call: CallbackQuery):
    from app.charts import generate_monthly_chart
    from aiogram.types import BufferedInputFile
    now = datetime.now()
    await call.answer("Генерирую график...")
    try:
        img = await generate_monthly_chart(call.from_user.id, now.year, now.month)
        await call.message.answer_photo(
            BufferedInputFile(img, filename="chart.png"),
            caption="График за " + str(now.month) + "/" + str(now.year),
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )
    except Exception as e:
        await call.message.answer("Ошибка генерации графика: " + str(e))


# --- Меню отчётов ---

@router.callback_query(F.data == "reports_menu")
async def cb_reports_menu(call: CallbackQuery):
    from app.keyboards import reports_menu_kb
    await call.message.edit_text(
        "Отчёты и инструменты:",
        parse_mode=None,
        reply_markup=reports_menu_kb()
    )


@router.callback_query(F.data.startswith("chart:"))
async def cb_chart_by_month(call: CallbackQuery):
    from app.charts import generate_monthly_chart
    from aiogram.types import BufferedInputFile
    parts = call.data.split(":")
    year, month = int(parts[1]), int(parts[2])
    MONTHS = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
              7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
    await call.answer("Генерирую график...")
    try:
        img = await generate_monthly_chart(call.from_user.id, year, month)
        await call.message.answer_photo(
            BufferedInputFile(img, filename="chart.png"),
            caption="График за " + MONTHS[month] + " " + str(year),
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ])
        )
    except Exception as e:
        await call.message.answer("Ошибка: " + str(e))


@router.callback_query(F.data == "export_all")
async def cb_export_all(call: CallbackQuery):
    from app.database import fetchall
    import openpyxl
    from io import BytesIO
    from aiogram.types import BufferedInputFile

    await call.answer("Готовлю выгрузку...")

    wb = openpyxl.Workbook()

    # 1. Транзакции
    ws1 = wb.active
    ws1.title = "Транзакции"
    ws1.append(["ID", "Сумма", "Тип", "Категория", "Комментарий", "Дата", "Создано"])
    rows = await fetchall("""
        SELECT t.id, t.amount, t.type, c.name as category,
               t.comment, t.transaction_date, t.created_at
        FROM transactions t
        LEFT JOIN categories c ON t.category_id = c.id
        WHERE t.user_id = %s
        ORDER BY t.transaction_date DESC
    """, (call.from_user.id,))
    for r in rows:
        ws1.append([r[0], float(r[1]), r[2], r[3], r[4], str(r[5]), str(r[6])])

    # 2. Категории
    ws2 = wb.create_sheet("Категории")
    ws2.append(["ID", "Название", "Тип"])
    cats = await fetchall("SELECT id, name, type FROM categories WHERE user_id = %s", (call.from_user.id,))
    for r in cats:
        ws2.append([r[0], r[1], r[2]])

    # 3. Заметки
    ws3 = wb.create_sheet("Заметки")
    ws3.append(["ID", "Текст", "Создано"])
    notes = await fetchall("SELECT id, text, created_at FROM notes WHERE user_id = %s ORDER BY created_at DESC", (call.from_user.id,))
    for r in notes:
        ws3.append([r[0], r[1], str(r[2])])

    # 4. Цели
    ws4 = wb.create_sheet("Цели")
    ws4.append(["ID", "Текст", "Создано"])
    goals = await fetchall("SELECT id, goal_text, created_at FROM user_goals WHERE user_id = %s ORDER BY created_at DESC", (call.from_user.id,))
    for r in goals:
        ws4.append([r[0], r[1], str(r[2])])

    # 5. История ИИ
    try:
        ws5 = wb.create_sheet("Диалоги ИИ")
        ws5.append(["ID", "Роль", "Сообщение", "Дата"])
        history = await fetchall("SELECT id, role, content, created_at FROM ai_history WHERE user_id = %s ORDER BY created_at", (call.from_user.id,))
        for r in history:
            ws5.append([r[0], r[1], r[2], str(r[3])])
    except Exception:
        pass

    # Сохраняем
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    await call.message.answer_document(
        BufferedInputFile(buf.read(), filename="export_" + str(call.from_user.id) + ".xlsx"),
        caption="Полная выгрузка базы данных",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ])
    )



HELP_TRIGGERS = [
    "помощь", "помоги", "не понимаю", "как пользоваться", "что умеешь",
    "как работает", "объясни", "расскажи", "хелп", "help", "мне нужна помощь",
    "что ты умеешь", "как добавить", "как записать", "как посмотреть",
    "не знаю", "подскажи", "инструкция"
]

SUPPORT_EMAIL = "findirvladislavku@gmail.com"
SUPPORT_CHANNEL = "@findir43"

BOT_KNOWLEDGE = """
Ты — встроенный помощник финансового бота "Баланс". Отвечай кратко, дружелюбно, с emoji.
Вот что умеет бот:

📥 РАСХОДЫ И ДОХОДЫ
- Кнопка "Ручной ввод" → выбери расход или доход → выбери категорию → напиши или скажи сумму с комментарием одним сообщением
- Быстрый ввод текстом: просто напиши сумму, например "500" или "-500 кофе"
- Голосом можно вносить операции из любого места бота: "кофе 250, такси 700", "получил зарплату 50000"
- Важно: в комментарии к голосовой операции лучше не говорить числа, количество и номера. Вместо "купил 2 кофе по 250" скажи "кофе 500", иначе бот может принять числа за отдельные операции

📋 ПОСЛЕДНИЕ ТРАНЗАКЦИИ
- Кнопка "Последние" → список последних операций
- Можно редактировать или удалять каждую транзакцию

📊 ОТЧЁТЫ
- Кнопка "Отчёты" → календарь, отчёт ДДС, графики по месяцам
- Выгрузка всей базы в Excel (на тарифе Премиум)

🤖 ИИ-ПОМОЩНИК
- Кнопка "ИИ-помощник" → финансовый советник на базе GPT-4o
- Можно обсудить бюджет, финансовый план, сложную ситуацию, долги, траты и доходы
- Можно установить или удалить финансовую цель
- На тарифе Премиум можно общаться вообще на любые темы, не только финансовые
- Голосовые сообщения работают внутри ассистента (кроме тарифа Скан и текст)

🧾 СКАНИРОВАНИЕ ЧЕКОВ
- Просто отправь фото чека → бот автоматически распознает итоговую сумму

📝 ЗАМЕТКИ
- Кнопка "Заметки" → текстовые заметки, не привязанные к транзакциям

🔁 РЕГУЛЯРНЫЕ ПЛАТЕЖИ
- Настрой регулярные платежи (аренда, подписки)
- В день платежа бот напомнит и спросит оплачено ли

📂 УПРАВЛЕНИЕ КАТЕГОРИЯМИ
- Чтобы заменить расходную статью: открой "Ручной ввод" → "Расход" и напиши "Заменить старую статью на новую"
- Чтобы заменить доходную статью: открой "Ручной ввод" → "Доход" и напиши "Заменить старую статью на новую"
- Там же можно написать "Добавить название статьи" или "Удалить название статьи"
- Эти команды можно сказать голосом, находясь внутри расходов или доходов

⭐ ТАРИФЫ
- Бесплатно: только просмотр последних и отчётов
- Скан и текст (149 руб/мес): текстовый ввод, чеки, 60 ИИ-сообщений, без голоса
- База (290 руб/мес): + голосовой ввод, 150 ИИ-сообщений
- Премиум (800 руб/мес): + безлимитный ИИ, свободные беседы на любые темы, Excel

📌 КОМАНДЫ
- /start — главное меню
- /help — эта справка
- /reset — удалить все транзакции
- /deleteaccount — удалить аккаунт полностью

❓ Если что-то непонятно — просто спроси меня! По сложным вопросам и проблемам пиши на """ + SUPPORT_EMAIL + """
"""


@router.message(Command("offer"))
async def cmd_offer(message: Message):
    from app.offer_text import OFFER_TEXT, OFFER_TEXT_2, OFFER_TEXT_3, OFFER_TEXT_4, OFFER_TEXT_5
    for part in [OFFER_TEXT, OFFER_TEXT_2, OFFER_TEXT_3, OFFER_TEXT_4, OFFER_TEXT_5]:
        await message.answer(part, parse_mode=None)
    await message.answer(
        "Меню",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ])
    )


@router.message(Command("help"))
async def cmd_help(message: Message):
    text = (
        "📌 Основные функции бота «Баланс»\n\n"
        "✍️ Ручной ввод — записать трату или поступление (кнопкой, текстом или голосом)\n"
        "📋 Последние — история операций, можно редактировать и удалять\n"
        "📊 Отчёты — календарь, отчёт ДДС, графики, выгрузка в Excel (Премиум)\n"
        "🤖 ИИ-помощник — бюджет, план, финансовые вопросы, цели; на Премиум — свободные беседы на любые темы\n"
        "🧾 Чеки — отправь фото чека, сумма распознается автоматически\n"
        "📝 Заметки — база фактов для ИИ: он видит заметки и использует их для отчётов, планов и анализа\n"
        "🔁 Регулярные платежи — календарь платежей с напоминаниями\n\n"
        "🎙 Голосовой ввод\n"
        "Операции можно вносить голосом из любого места бота.\n"
        "Примеры: \"кофе 250, такси 700\", \"получил зарплату 50000\", \"продукты 3200 и аренда 30000\".\n"
        "Важно: в комментарии не указывай числа, количество и номера. Лучше сказать \"кофе 500\", а не \"2 кофе по 250\", чтобы бот не принял числа за отдельные операции.\n\n"
        "📂 Категории\n"
        "Чтобы изменить расходную статью: открой \"Ручной ввод\" → \"Расход\".\n"
        "Чтобы изменить доходную статью: открой \"Ручной ввод\" → \"Доход\".\n"
        "Внутри нужного раздела можно написать или сказать голосом:\n"
        "— \"Заменить Кафе на Рестораны\"\n"
        "— \"Добавить Подарки\"\n"
        "— \"Удалить Старую статью\"\n\n"
        "🤖 ИИ-помощник\n"
        "В ИИ-помощнике можно обсудить бюджет, финансовый план, проблемы с расходами, идеи экономии, а также установить или удалить финансовую цель.\n\n"
        "⭐ Тарифы\n"
        "Скан и текст — 149 руб/мес, текст + чеки, 60 ИИ-сообщений\n"
        "База — 290 руб/мес, + голос, 150 ИИ-сообщений\n"
        "Премиум — 800 руб/мес, безлимитный ИИ + свободные беседы + Excel\n\n"
        "📌 Команды\n"
        "/start — главное меню\n"
        "/help — эта справка\n"
        "/offer — публичная оферта (условия использования)\n"
        "/reset — удалить все транзакции\n"
        "/deleteaccount — удалить аккаунт полностью\n\n"
        "✉️ Связь и поддержка\n"
        "Если возникла проблема или есть вопрос — пиши на " + SUPPORT_EMAIL + "\n"
        "Канал с обучающими материалами: " + SUPPORT_CHANNEL
    )
    await message.answer(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
    ]))


async def answer_help_question(user_text: str, bot_answer_func=None):
    """Отвечает на вопрос о работе бота через GPT"""
    import httpx, os
    async with httpx.AsyncClient() as client:
        response = await client.post(
            "https://api.openai.com/v1/chat/completions",
            headers={
                "Authorization": "Bearer " + os.environ.get("OPENAI_API_KEY", ""),
                "Content-Type": "application/json",
            },
            json={
                "model": "gpt-4o",
                "max_tokens": 400,
                "messages": [
                    {"role": "system", "content": BOT_KNOWLEDGE},
                    {"role": "user", "content": user_text}
                ],
            },
            timeout=15.0
        )
        data = response.json()
        return data["choices"][0]["message"]["content"]


def _format_tx_preview(tx, public_id: str | None = None) -> str:
    tx_id, tx_date, amount, type_, comment, category = tx
    sign = "-" if type_ == "expense" else "+"
    comment_part = _format_comment_inline(comment)
    label = public_id or str(tx_id)
    return f"#{label} {tx_date.strftime('%d.%m')} {sign}{abs(float(amount)):,.0f} {category or ''}{comment_part}"


async def _format_tx_preview_async(user_id: int, tx) -> str:
    return _format_tx_preview(tx, await public_number_for_transaction(user_id, int(tx[0])))


async def _format_tx_preview_lines(user_id: int, txs) -> str:
    tx_list = list(txs)
    numbers = await public_numbers_for_ids(user_id, [int(tx[0]) for tx in tx_list])
    return "\n".join(
        _format_tx_preview(tx, numbers.get(int(tx[0]), str(tx[0])))
        for tx in tx_list
    )


async def _format_recent_tx_lines(user_id: int, txs, limit: int = 8) -> str:
    tx_list = list(txs[:limit])
    numbers = await public_numbers_for_ids(user_id, [int(tx["id"]) for tx in tx_list])
    lines = []
    for tx in tx_list:
        sign = "-" if tx["type"] == "expense" else "+"
        date_value = tx["transaction_date"].strftime("%d.%m")
        comment = _format_comment_inline(tx.get("comment"))
        public_id = numbers.get(int(tx["id"]), str(tx["id"]))
        lines.append(
            f"#{public_id} {date_value} {sign}{abs(float(tx['amount'])):,.0f} "
            f"{tx.get('category_name') or ''}{comment}"
        )
    return "\n".join(lines)


async def _answer_delete_needs_choice(message: Message, user_id: int, prefix: str):
    txs = await get_recent_transactions(user_id, limit=8)
    text = prefix
    if txs:
        text += "\n\nПоследние операции:\n" + await _format_recent_tx_lines(user_id, txs)
        text += "\n\nНажми «Удалить по номеру» и отправь номер нужной операции."
    else:
        text += "\n\nПока нет операций для удаления."
    await message.answer(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Удалить по номеру", callback_data="delete_by_id")],
            [InlineKeyboardButton(text="Открыть список операций", callback_data="recent")],
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ]),
    )


async def handle_intent_message(message: Message, state: FSMContext, text: str, source: str = "text") -> bool:
    from app.database import (
        get_last_transaction,
        get_transaction_by_id,
        delete_transaction_by_id,
    )
    from app.services.category_commands import apply_category_command
    from app.services.intent_router import parse_user_intent

    intent = await parse_user_intent(message.from_user.id, text, source)
    name = intent.get("intent")
    params = intent.get("params", {})

    if name == "unknown":
        return False

    if name == "open_main_menu":
        await state.clear()
        await message.answer("Выбирай действие:", reply_markup=main_menu())
        return True

    if name == "show_report":
        report_text, kb = await render_month_report(
            message.from_user.id,
            int(params["year"]),
            int(params["month"]),
        )
        await message.answer(report_text, parse_mode="HTML", reply_markup=kb)
        return True

    if name == "beautiful_report":
        from app.services.kie_reports import describe_report_plan, is_default_beautiful_report_request

        user_prompt = params.get("user_prompt") or text
        year = int(params["year"])
        month = int(params["month"])
        if is_default_beautiful_report_request(user_prompt):
            await _start_beautiful_report_generation(
                message,
                state,
                message.from_user.id,
                year,
                month,
                user_prompt,
            )
            return True

        await state.set_state(BeautifulReportState.waiting_confirm)
        await state.update_data(beautiful_report={
            "year": year,
            "month": month,
            "user_prompt": user_prompt,
        })
        await message.answer(
            describe_report_plan(year, month, user_prompt),
            reply_markup=_beautiful_report_confirm_keyboard(),
        )
        return True

    if name == "show_recent":
        txs = await get_recent_transactions(message.from_user.id, limit=10)
        if not txs:
            await message.answer(
                "Пока нет транзакций.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True
        await message.answer(
            "Последние транзакции:\n\n" + await _format_recent_tx_lines(message.from_user.id, txs, limit=10),
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Открыть список", callback_data="recent")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return True

    if name == "show_calendar":
        await message.answer(
            "Открываю платёжный календарь.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="🗓 Календарь", callback_data="calendar")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return True

    if name == "add_recurring_payment":
        from app.database import add_recurring_payment, can_use_feature
        from app.services.recurring_commands import parse_recurring_add

        if not await can_use_feature(message.from_user.id, "calendar"):
            await message.answer(
                "Платёжный календарь доступен с тарифа Старт и выше.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Тарифы", callback_data="premium")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

        parsed = parse_recurring_add(params.get("text") or text)
        if not parsed:
            await message.answer(
                "Не понял платёж. Напиши, например: «добавь аренду 30000 каждый месяц 5 числа».",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Открыть календарь", callback_data="calendar")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

        categories = await get_categories(message.from_user.id, type_="expense")
        cat_id = categories[0]["id"] if categories else None
        await add_recurring_payment(
            user_id=message.from_user.id,
            name=parsed["name"],
            amount=parsed["amount"],
            type_="expense",
            kind="fixed",
            category_id=cat_id,
            repeat_type=parsed["repeat_type"],
            repeat_day_of_month=parsed.get("repeat_day_month"),
            repeat_day_of_week=parsed.get("repeat_day_week"),
            remind_days_before=1,
            amount_is_approximate=parsed.get("is_approx", False),
        )
        if parsed["repeat_type"] == "monthly":
            repeat_text = f"каждый месяц {parsed.get('repeat_day_month') or 1} числа"
        else:
            repeat_text = "еженедельно"
        await message.answer(
            f"Добавил в платёжный календарь:\n"
            f"{parsed['name']} — {parsed['amount']:,.0f} ₽, {repeat_text}.\n\n"
            "Теперь буду напоминать об этом платеже заранее.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Открыть календарь", callback_data="calendar")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return True

    if name == "delete_recurring_payment":
        from app.services.recurring_commands import find_recurring_payment

        payment, variants = await find_recurring_payment(message.from_user.id, params.get("text") or text)
        if not payment:
            if variants:
                lines = []
                for item in variants:
                    lines.append(f"#{item[0]} {item[2]} — {float(item[3] or 0):,.0f} ₽")
                await message.answer(
                    "Нашёл несколько похожих платежей:\n"
                    + "\n".join(lines)
                    + "\n\nОткрой календарь и выбери нужный платёж для удаления.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Открыть календарь", callback_data="calendar")],
                        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                    ]),
                )
                return True
            await message.answer(
                "Не нашёл такой платёж в календаре.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Открыть календарь", callback_data="calendar")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

        await message.answer(
            f"Удалить платёж из календаря?\n\n{payment[2]} — {float(payment[3] or 0):,.0f} ₽",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="Да, удалить", callback_data=f"confirm_delete_recurring_intent:{payment[0]}"),
                    InlineKeyboardButton(text="Отмена", callback_data="calendar"),
                ],
            ]),
        )
        return True

    if name == "edit_recurring_payment":
        from app.database import can_use_feature
        from app.services.recurring_commands import find_recurring_payment, parse_recurring_edit

        if not await can_use_feature(message.from_user.id, "calendar"):
            await message.answer(
                "Платёжный календарь доступен с тарифа Старт и выше.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Тарифы", callback_data="premium")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

        raw_text = params.get("text") or text
        payment, variants = await find_recurring_payment(message.from_user.id, raw_text)
        if not payment:
            if variants:
                lines = [f"#{item[0]} {item[2]} — {float(item[3] or 0):,.0f} ₽" for item in variants]
                await message.answer(
                    "Нашёл несколько похожих платежей:\n"
                    + "\n".join(lines)
                    + "\n\nОткрой календарь и выбери нужный платёж.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Открыть календарь", callback_data="calendar")],
                        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                    ]),
                )
                return True
            await message.answer(
                "Не нашёл такой платёж в календаре. Можно открыть календарь и выбрать его вручную.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Открыть календарь", callback_data="calendar")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

        parsed = parse_recurring_edit(raw_text)
        if not parsed:
            await message.answer(
                "Понял платёж, но не понял, что поменять. Например: "
                "«перенеси аренду на 10 число» или «измени кредит на 25000 рублей».",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Открыть календарь", callback_data="calendar")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

        repeat_type = parsed.get("repeat_type") or payment[8]
        repeat_day_month = payment[9] if repeat_type == "monthly" else None
        repeat_day_week = payment[10] if repeat_type == "weekly" else None
        if parsed.get("repeat_day_month") is not None:
            repeat_day_month = parsed["repeat_day_month"]
            repeat_day_week = None
        if parsed.get("repeat_day_week") is not None:
            repeat_day_week = parsed["repeat_day_week"]
            repeat_day_month = None

        edit = {
            "payment_id": payment[0],
            "amount": parsed["amount"] if parsed.get("amount") is not None else float(payment[3] or 0),
            "amount_is_approximate": (
                parsed["is_approx"] if parsed.get("is_approx") is not None else bool(payment[4])
            ),
            "repeat_type": repeat_type,
            "repeat_day_of_month": repeat_day_month,
            "repeat_day_of_week": repeat_day_week,
            "next_trigger_date": _next_recurring_date(repeat_type, repeat_day_month, repeat_day_week),
        }
        updated_payment = list(payment)
        updated_payment[3] = edit["amount"]
        updated_payment[4] = edit["amount_is_approximate"]
        updated_payment[8] = edit["repeat_type"]
        updated_payment[9] = edit["repeat_day_of_month"]
        updated_payment[10] = edit["repeat_day_of_week"]

        await state.update_data(pending_recurring_edit=edit)
        await message.answer(
            "Изменить платёж в календаре?\n\n"
            f"Было: {_format_recurring_payment(payment)}\n"
            f"Будет: {_format_recurring_payment(updated_payment)}",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="Да, изменить", callback_data="confirm_edit_recurring_intent"),
                    InlineKeyboardButton(text="Отмена", callback_data="cancel_edit_recurring_intent"),
                ],
                [InlineKeyboardButton(text="Открыть календарь", callback_data="calendar")],
            ]),
        )
        return True

    if name == "clarify_category_or_transaction_edit":
        await state.update_data(action_pending_text=params.get("text") or text)
        await message.answer(
            "Что меняем?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Статьи расходов/доходов", callback_data="clarify_edit_categories")],
                [InlineKeyboardButton(text="Записанные операции", callback_data="clarify_edit_transactions")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return True

    if name == "delete_transaction":
        tx = None
        variants = []
        tx_id = params.get("tx_id")
        if tx_id:
            resolved_tx_id = await resolve_transaction_reference(message.from_user.id, tx_id)
            if resolved_tx_id:
                tx = await get_transaction_by_id(message.from_user.id, resolved_tx_id)
        elif params.get("last"):
            tx = await get_last_transaction(message.from_user.id)
        if not tx and not params.get("last"):
            from app.services.transaction_commands import find_transaction_from_text
            tx, variants = await find_transaction_from_text(message.from_user.id, params.get("text") or text)

        if not tx:
            if variants:
                await message.answer(
                    "Нашёл несколько похожих транзакций:\n"
                    + await _format_tx_preview_lines(message.from_user.id, variants)
                    + "\n\nНапиши номер нужной транзакции.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Удалить по номеру", callback_data="delete_by_id")],
                        [InlineKeyboardButton(text="Открыть список операций", callback_data="recent")],
                        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                    ]),
                )
                return True
            await _answer_delete_needs_choice(
                message,
                message.from_user.id,
                "Не понял, какую операцию удалить.",
            )
            return True

        await message.answer(
            "Вы имеете в виду эту транзакцию?\n" + await _format_tx_preview_async(message.from_user.id, tx),
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="Да, удалить", callback_data=f"confirm_delete_intent:{tx[0]}"),
                    InlineKeyboardButton(text="Не то, список", callback_data="recent"),
                ],
                [InlineKeyboardButton(text="Удалить по номеру", callback_data="delete_by_id")],
            ]),
        )
        return True

    if name == "edit_transaction":
        from app.services.transaction_commands import find_transaction_for_edit

        tx, variants = await find_transaction_for_edit(message.from_user.id, params.get("text") or text)
        if not tx:
            if variants:
                await message.answer(
                    "Нашёл несколько похожих операций:\n"
                    + await _format_tx_preview_lines(message.from_user.id, variants)
                    + "\n\nВыбери операцию из списка или напиши её номер и правку.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Открыть список операций", callback_data="recent")],
                        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                    ]),
                )
                return True
            txs = await get_recent_transactions(message.from_user.id, limit=8)
            text_for_user = "Не понял, какую операцию изменить."
            if txs:
                text_for_user += "\n\nПоследние операции:\n" + await _format_recent_tx_lines(message.from_user.id, txs)
                text_for_user += "\n\nОткрой список, выбери операцию и напиши правку обычными словами."
            await message.answer(
                text_for_user,
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Открыть список операций", callback_data="recent")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

        current, updated = await _build_transaction_edit(message.from_user.id, params.get("text") or text, tx[0])
        if not current:
            await message.answer("Операция не найдена.", reply_markup=main_menu())
            return True
        if not updated:
            await message.answer(
                "Операцию нашёл, но не понял, что именно изменить. Напиши проще: «измени дату на 05.07» или «исправь сумму на 1023 руб».",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Открыть список операций", callback_data="recent")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

        await state.update_data(pending_transaction_edit=updated)
        await message.answer(
            "Изменить операцию?\n\n"
            + "Было: " + _format_current_tx(current) + "\n"
            + "Будет: " + _format_current_tx(updated),
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(text="Да, изменить", callback_data="confirm_edit_transaction"),
                    InlineKeyboardButton(text="Отмена", callback_data="cancel_edit_transaction"),
                ],
                [InlineKeyboardButton(text="Открыть список операций", callback_data="recent")],
            ]),
        )
        return True

    if name == "change_transaction_category":
        from app.services.transaction_commands import (
            find_transaction_for_category_change,
            parse_category_change,
        )
        parsed = await parse_category_change(message.from_user.id, params.get("text") or text)
        category_id = parsed.get("new_category_id")
        category_name = parsed.get("new_category_name")
        if not category_id:
            await message.answer(
                "Не понял новую категорию. Напиши, например: «поменяй категорию транзакции #712 на Транспорт».",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True
        tx, variants = await find_transaction_for_category_change(
            message.from_user.id,
            params.get("text") or text,
            parsed,
        )
        if not tx:
            if variants:
                await message.answer(
                    "Нашёл несколько похожих транзакций:\n"
                    + await _format_tx_preview_lines(message.from_user.id, variants)
                    + "\n\nНапиши номер транзакции и новую категорию точнее.",
                    reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                        [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                    ]),
                )
                return True
            await message.answer(
                "Не нашёл транзакцию для смены категории. Укажи номер или сумму/месяц точнее.",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True
        await message.answer(
            "Вы имеете в виду эту транзакцию?\n"
            + await _format_tx_preview_async(message.from_user.id, tx)
            + f"\n\nПоменять категорию на «{category_name}»?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="Да, поменять",
                        callback_data=f"confirm_change_tx_category:{tx[0]}:{category_id}",
                    ),
                    InlineKeyboardButton(text="Нет", callback_data="cancel_change_tx_category"),
                ],
            ]),
        )
        return True

    if name in ("rename_category", "add_category", "delete_category", "set_category_kind"):
        scope_type = params.get("scope_type", "expense")
        result = await apply_category_command(message.from_user.id, params.get("command"), scope_type)
        await message.answer(
            result,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Ручной ввод", callback_data="manual_input")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return True

    if name == "add_transaction":
        added = []
        saved_ids = []
        saved_items = []
        for tx in params.get("transactions", []):
            saved = await create_transaction(
                user_id=message.from_user.id,
                category_id=tx["category_id"],
                amount=tx["amount"],
                type_=tx["type"],
                kind=tx.get("kind"),
                comment=tx.get("comment") or "",
                transaction_date=tx.get("transaction_date"),
                pnl_period=tx.get("pnl_period"),
            )
            saved_ids.append(saved["id"])
            saved_items.append((saved["id"], tx))

        public_ids = await public_numbers_for_ids(message.from_user.id, saved_ids)
        for saved_id, tx in saved_items:
            sign = "-" if tx["type"] == "expense" else "+"
            public_id = public_ids.get(int(saved_id), str(saved_id))
            added.append(
                f"{sign}{int(float(tx['amount']))} руб. — {tx['category_name']} #{public_id}"
                + _format_comment_inline(tx.get("comment"))
            )

        if added:
            response_text = "Записано:\n" + "\n".join(added)
            insight = await build_first_transaction_insight(message.from_user.id, saved_ids)
            if insight:
                response_text += "\n\n" + insight
            await message.answer(
                response_text,
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return True

    return False


async def send_text_to_ai_assistant(message: Message, state: FSMContext, user_id: int, user_text: str):
    from app.database import get_user_tier
    from app.handlers.ai_assistant import (
        AI_MODE_TALK,
        AIState,
        _ai_mode_keyboard,
        check_ai_limit,
        get_ai_response,
        log_ai_usage,
        process_actions,
    )

    tier = await get_user_tier(user_id)
    if tier == "free":
        await message.answer(
            "ИИ-помощник доступен с тарифа Старт.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Тарифы", callback_data="premium")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    can, used, limit = await check_ai_limit(user_id)
    if not can:
        await message.answer(
            "Лимит ИИ-помощника исчерпан на этот месяц.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Тарифы", callback_data="premium")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    await state.set_state(AIState.chatting)
    thinking = await message.answer("Передаю в ИИ-помощник...")
    try:
        ai_text, new_history = await get_ai_response(user_id, user_text, [], tier=tier, mode=AI_MODE_TALK)
        await state.update_data(history=new_history[-20:], ai_mode=AI_MODE_TALK)
        await log_ai_usage(user_id)
        clean_text, actions_log = await process_actions(
            user_id,
            ai_text,
            user_text,
            allow_actions=False,
        )
        limit_str = "безлимит" if limit >= 9999 else str(used + 1) + "/" + str(limit)
        await thinking.edit_text(
            "🗣️ " + clean_text + actions_log + "\n\n[" + limit_str + "]",
            parse_mode=None,
            reply_markup=_ai_mode_keyboard(AI_MODE_TALK),
        )
    except Exception as e:
        await thinking.edit_text(
            "Не смог передать в ИИ-помощник: " + str(e),
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )


@router.callback_query(F.data == "ask_ai_pending")
async def cb_ask_ai_pending(call: CallbackQuery, state: FSMContext):
    await call.answer()
    data = await state.get_data()
    user_text = data.get("ai_pending_text")
    if not user_text:
        await call.message.answer(
            "Не нашёл текст для ИИ. Напиши вопрос ещё раз.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="ИИ-помощник", callback_data="ai_assistant")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return
    await send_text_to_ai_assistant(call.message, state, call.from_user.id, user_text)


@router.callback_query(F.data == "clarify_edit_categories")
async def cb_clarify_edit_categories(call: CallbackQuery, state: FSMContext):
    await call.answer()
    data = await state.get_data()
    user_text = data.get("action_pending_text")
    if not user_text:
        await call.message.answer(
            "Не нашёл фразу для настройки статей. Напиши её ещё раз.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    from app.services.category_commands import parse_category_command, apply_category_command

    result_text = None
    for scope_type in ("expense", "income"):
        command = await parse_category_command(call.from_user.id, user_text, scope_type)
        if command and command.get("intent") != "unknown":
            result = await apply_category_command(call.from_user.id, command, scope_type)
            if "не найдена" not in result.lower():
                result_text = result
                break
            result_text = result_text or result

    if result_text:
        await call.message.answer(
            result_text,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Показать категории", callback_data="categories_list")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
    else:
        await call.message.answer(
            "Не понял, какую статью заменить. Напиши так: «замени категорию Здоровье на Развитие».",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Показать категории", callback_data="categories_list")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )


@router.callback_query(F.data == "clarify_edit_transactions")
async def cb_clarify_edit_transactions(call: CallbackQuery, state: FSMContext):
    await call.answer()
    await call.message.answer(
        "Выбери операцию из списка, потом напиши правку обычными словами.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="Открыть операции", callback_data="recent")],
            [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
        ]),
    )


@router.message(F.text & ~F.text.startswith("/"), StateFilter(default_state))
async def msg_free_text(message: Message, state: FSMContext):
    # ИИ-хелпер — триггер по ключевым фразам
    text_lower = message.text.lower().strip()
    if _is_no_spend_reply(message.text):
        await message.answer(
            "Хорошо, тогда можно записать позже, когда появятся расходы.\n\n"
            "Например: «500 на продукты» или «кофе 250».",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    if any(trigger in text_lower for trigger in HELP_TRIGGERS):
        thinking = await message.answer("Сейчас расскажу...")
        try:
            answer = await answer_help_question(message.text)
            await thinking.delete()
            await message.answer(
                "❓ " + answer,
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ])
            )
        except Exception as e:
            await thinking.delete()
            await message.answer("Ошибка хелпера: " + str(e))
        return

    handled = await handle_intent_message(message, state, message.text, source="text")
    if not handled:
        if _looks_like_edit_confusion(message.text):
            await state.update_data(action_pending_text=message.text)
            await message.answer(
                "Что меняем?",
                reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                    [InlineKeyboardButton(text="Статьи расходов/доходов", callback_data="clarify_edit_categories")],
                    [InlineKeyboardButton(text="Записанные операции", callback_data="clarify_edit_transactions")],
                    [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
                ]),
            )
            return
        await state.update_data(ai_pending_text=message.text)
        await message.answer(
            "Не понял это как транзакцию или команду. Могу передать фразу в ИИ-помощник.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Спросить ИИ", callback_data="ask_ai_pending")],
                [InlineKeyboardButton(text="Открыть ИИ-помощник", callback_data="ai_assistant")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return


@router.message(Command("category"))
async def cmd_category(message: Message):
    from app.database import get_categories, execute
    from app.services.category_commands import normalize_category_display_name
    args = message.text.strip()[len("/category"):].strip()

    if not args or args == "list":
        await message.answer(
            await render_categories_text(message.from_user.id),
            parse_mode=None,
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="Быстрая смена статей", callback_data="quick_category_rename")],
                [InlineKeyboardButton(text="Меню", callback_data="main_menu")],
            ]),
        )
        return

    # rename
    if args.startswith("rename"):
        import re
        m = re.findall(r'"([^"]+)"', args)
        if len(m) < 2:
            await message.answer('Формат: /category rename "Старое" "Новое"')
            return
        old_name, new_name = m[0], normalize_category_display_name(m[1])
        result = await execute(
            "UPDATE categories SET name=%s WHERE name=%s AND user_id=%s",
            (new_name, old_name, message.from_user.id)
        )
        await message.answer(f'Категория "{old_name}" → "{new_name}"' if result else f'Категория "{old_name}" не найдена.')
        return

    # add
    if args.startswith("add"):
        import re
        m = re.findall(r'"([^"]+)"', args)
        type_ = 'expense' if 'income' not in args else 'income'
        if not m:
            await message.answer('Формат: /category add "Название" expense')
            return
        name = normalize_category_display_name(m[0])
        kind = "income" if type_ == "income" else "variable"
        await execute(
            "INSERT INTO categories (user_id, name, type, kind) VALUES (%s, %s, %s, %s) ON CONFLICT DO NOTHING",
            (message.from_user.id, name, type_, kind)
        )
        type_ru = "доход" if type_ == 'income' else "расход"
        await message.answer(f'Категория "{name}" ({type_ru}) добавлена.')
        return

    # delete
    if args.startswith("delete"):
        import re
        m = re.findall(r'"([^"]+)"', args)
        if not m:
            await message.answer('Формат: /category delete "Название"')
            return
        name = m[0]
        await execute(
            "DELETE FROM categories WHERE name=%s AND user_id=%s",
            (name, message.from_user.id)
        )
        await message.answer(f'Категория "{name}" удалена.')
        return

    await message.answer("Неизвестная команда. Используй: list, rename, add, delete")
